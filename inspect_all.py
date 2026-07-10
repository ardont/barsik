import openpyxl

file_path = r"C:\Users\Maxim\Downloads\08.07_1.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb["Лист2"]

with open("sheet2_all.txt", "w", encoding="utf-8") as f:
    f.write("Row | TicketProf (A-G) | Bars Tour (H-M)\n")
    f.write("-" * 120 + "\n")
    for r in range(2, ws.max_row + 1):
        # TicketProf cols: A (1) to G (7)
        tp_vals = []
        for c in range(1, 8):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                tp_vals.append(f"{openpyxl.utils.get_column_letter(c)}:{v}")
        tp_str = ", ".join(tp_vals)
        
        # Bars Tour cols: H (8) to M (13)
        bt_vals = []
        for c in range(8, 14):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                bt_vals.append(f"{openpyxl.utils.get_column_letter(c)}:{v}")
        bt_str = ", ".join(bt_vals)
        
        f.write(f"{r:3d} | {tp_str:<55} | {bt_str}\n")
wb.close()
print("Done")
