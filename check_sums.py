import openpyxl

file_path = r"C:\Users\Maxim\Downloads\08.07_1.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb["Лист2"]

current_doc = None
sub_sum = 0
header_row = None
header_val = None

print("Checking header vs sub-item sum consistency:")
for r in range(2, ws.max_row + 1):
    val_a = ws.cell(row=r, column=1).value
    val_b = ws.cell(row=r, column=2).value
    val_c = ws.cell(row=r, column=3).value
    val_d = ws.cell(row=r, column=4).value
    
    is_header = False
    if val_a is not None and val_b is not None:
        if any(kw in str(val_b) for kw in ["Продажа", "Оплата", "Возврат", "Корректировка"]):
            is_header = True
            
    if is_header:
        if current_doc:
            print(f"Doc Row {header_row} | {current_doc[:35]:<35} | Header sum: {header_val:<10} | Sub sum: {sub_sum:<10} | Match: {abs(header_val - sub_sum) < 0.01}")
        current_doc = val_b
        header_row = r
        header_val = val_c if val_c is not None else (-val_d if val_d is not None else 0.0)
        sub_sum = 0
    elif val_a is not None and val_b is None:
        amt = val_c if val_c is not None else (-val_d if val_d is not None else 0.0)
        sub_sum += amt

if current_doc:
    print(f"Doc Row {header_row} | {current_doc[:35]:<35} | Header sum: {header_val:<10} | Sub sum: {sub_sum:<10} | Match: {abs(header_val - sub_sum) < 0.01}")

wb.close()
