import openpyxl
import json
import sys

def inspect_file(filepath, out_name):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    with open(out_name, 'w', encoding='utf-8') as f:
        f.write(f"FILE: {filepath}\n")
        f.write(f"Sheets: {wb.sheetnames}\n\n")
        
        for sname in wb.sheetnames:
            sheet = wb[sname]
            f.write(f"--- SHEET: {sname} (max_row={sheet.max_row}, max_col={sheet.max_column}) ---\n")
            for r in range(1, sheet.max_row + 1):
                row_vals = [sheet.cell(r, c).value for c in range(1, sheet.max_column + 1)]
                # check cell fills / comments / font colors
                highlights = []
                for c in range(1, sheet.max_column + 1):
                    cell = sheet.cell(r, c)
                    fill = cell.fill.start_color.rgb if cell.fill and cell.fill.start_color else None
                    comment = cell.comment.text if cell.comment else None
                    font_color = cell.font.color.rgb if cell.font and cell.font.color else None
                    if fill or comment or font_color:
                        highlights.append((c, fill, font_color, comment))
                
                if any(v is not None for v in row_vals) or highlights:
                    f.write(f"Row {r:3d}: {row_vals} | Highlights: {highlights}\n")

if __name__ == '__main__':
    inspect_file(r'C:\Users\Maxim\Downloads\13-19.07.xlsx', 'inspect_13_19_orig.txt')
    inspect_file(r'C:\Users\Maxim\Downloads\13-19.07_сопоставлено.xlsx', 'inspect_13_19_sopostavleno.txt')
    print("Done inspecting.")
