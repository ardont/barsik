import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\Maxim\Downloads\13-19.07_сопоставлено.xlsx', data_only=True)

with open('sop_sheet_dump.txt', 'w', encoding='utf-8') as f:
    for sname in wb.sheetnames:
        ws = wb[sname]
        f.write(f"\n================ SHEET: {sname} (max_row={ws.max_row}, max_col={ws.max_column}) ================\n")
        for r in range(1, ws.max_row + 1):
            row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            if any(v is not None for v in row_vals):
                f.write(f"Row {r:2d}: {row_vals}\n")

print("Dump created.")
