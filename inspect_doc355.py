import openpyxl

wb = openpyxl.load_workbook(r"C:\Users\Maxim\Downloads\20-26_07_26.xlsx", data_only=True)
ws = wb.active

for r in range(55, 78):
    c1_4 = [ws.cell(r, c).value for c in range(1, 5)]
    c8_11 = [ws.cell(r, c).value for c in range(8, 12)]
    print(f"Row {r:3d} | TP (A-D): {c1_4} | BT (H-K): {c8_11}")
