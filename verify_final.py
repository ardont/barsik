import openpyxl
import sys
sys.stdout.reconfigure(encoding='utf-8')
from engine.loader import load_data
from engine.matcher import match_records
from engine.calculator import calculate_reconciliation
from reports.excel_export import export_to_excel
from engine.normalizer import get_primary_id

tp_items, bt_items = load_data(r'C:\Users\Maxim\Downloads\13-19.07.xlsx')
matches, unmatched_tp, unmatched_bt = match_records(tp_items, bt_items, manual_links={}, settings={'simple_mode': True})
summary = calculate_reconciliation(tp_items, bt_items, matches)

output_excel = r'C:\Users\Maxim\Downloads\13-19.07_fixed_report.xlsx'
export_to_excel(tp_items, bt_items, matches, unmatched_tp, unmatched_bt, summary, output_excel)

print(f"Exported to {output_excel}")

wb = openpyxl.load_workbook(output_excel, data_only=True)
ws_all = wb["📊 Все сопоставления"]

print("\n--- FIRST 15 ROWS OF '📊 Все сопоставления' ---")
for r in range(1, 16):
    row_vals = [ws_all.cell(r, c).value for c in range(1, 12)]
    print(f"Row {r:2d}: {row_vals}")

print("\n--- UNMATCHED ROWS ON '📊 Все сопоставления' ---")
for r in range(25, 30):
    row_vals = [ws_all.cell(r, c).value for c in range(1, 12)]
    print(f"Row {r:2d}: {row_vals}")

print("\n--- SHEET '⚠ Несоответствия' ---")
ws_m = wb["⚠ Несоответствия"]
for r in range(1, ws_m.max_row + 1):
    row_vals = [ws_m.cell(r, c).value for c in range(1, 12)]
    print(f"Row {r:2d}: {row_vals}")
