import openpyxl

def find_marks(path, out):
    wb = openpyxl.load_workbook(path, data_only=False)
    out.write("==================================================\n")
    out.write(f"FILE: {path}\n")
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        out.write(f"\n--- Sheet: {sheetname} (rows={ws.max_row}, cols={ws.max_column}) ---\n")
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(r, c)
                val = cell.value
                comment = cell.comment.text if cell.comment else None
                
                fill = None
                if cell.fill and cell.fill.fill_type:
                    fg = cell.fill.fgColor
                    if fg:
                        fill = fg.rgb or fg.theme or fg.indexed
                
                font = None
                if cell.font and cell.font.color:
                    font = cell.font.color.rgb or cell.font.color.theme
                
                s_fill = str(fill) if fill is not None else ""
                s_font = str(font) if font is not None else ""
                
                # Check for standard colors vs manual colors
                # Standard colors in sopostavleno:
                # E2EFDA (green match), FFC7CE (red discrepancy), FCE4D6 (orange unmatched), 1F497D (blue header)
                if s_fill not in ['', 'None', '00000000', 'FFFFFFFF', '00FFFFFF', 'FFE2EFDA', 'FFFFC7CE', 'FFFCE4D6', 'FF1F497D'] or \
                   s_font not in ['', 'None', '00000000', 'FF000000', 'FFFFFFFF', '00FFFFFF', 'FF1F497D'] or comment:
                    out.write(f"Row {r:2d}, Col {c:2d} ({cell.coordinate}): val={repr(val)} | fill={s_fill} | font={s_font} | comment={repr(comment)}\n")

with open('user_marks_report.txt', 'w', encoding='utf-8') as out:
    find_marks(r'C:\Users\Maxim\Downloads\13-19.07.xlsx', out)
    find_marks(r'C:\Users\Maxim\Downloads\13-19.07_сопоставлено.xlsx', out)

print("Done.")
