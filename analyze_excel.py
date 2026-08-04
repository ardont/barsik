import openpyxl

def analyze_excel(path, report_path):
    wb = openpyxl.load_workbook(path, data_only=True)
    with open(report_path, 'w', encoding='utf-8') as out:
        out.write(f"FILE: {path}\n")
        for sname in wb.sheetnames:
            ws = wb[sname]
            out.write(f"\n--- SHEET: {sname} ---\n")
            for r in range(1, ws.max_row + 1):
                row_items = []
                for c in range(1, ws.max_column + 1):
                    cell = ws.cell(r, c)
                    val = cell.value
                    fill_color = None
                    if cell.fill and cell.fill.fill_type:
                        fg = cell.fill.fgColor
                        if fg:
                            fill_color = fg.rgb or fg.theme or fg.indexed
                    
                    font_color = None
                    if cell.font and cell.font.color:
                        font_color = cell.font.color.rgb or cell.font.color.theme
                    
                    comment = cell.comment.text if cell.comment else None
                    
                    row_items.append({
                        'c': c, 'val': val, 'fill': fill_color, 'font': font_color, 'comment': comment
                    })
                
                # Check if there are fills or comments or font colors worth reporting
                interesting = [item for item in row_items if item['fill'] or item['comment'] or (item['font'] and item['font'] not in ['00000000', 'FF000000', 'FFFFFFFF', 0, 1])]
                if interesting or any(item['val'] is not None for item in row_items):
                    out.write(f"Row {r:2d}:\n")
                    for item in row_items:
                        if item['val'] is not None or item['fill'] or item['comment']:
                            out.write(f"  Col {item['c']:2d}: val={repr(item['val'])} fill={item['fill']} font={item['font']} comment={repr(item['comment'])}\n")

analyze_excel(r'C:\Users\Maxim\Downloads\13-19.07.xlsx', 'report_orig.txt')
analyze_excel(r'C:\Users\Maxim\Downloads\13-19.07_сопоставлено.xlsx', 'report_sopostavleno.txt')
print("Reports created.")
