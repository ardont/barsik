import openpyxl

file_path = r"C:\Users\Maxim\Downloads\08.07_1.xlsx"

wb = openpyxl.load_workbook(file_path, data_only=True)

for name in wb.sheetnames:
    ws = wb[name]
    print(f"Sheet '{name}' dimensions: {ws.dimensions}")
    # Let's count how many rows have non-empty columns A and H
    a_rows = 0
    h_rows = 0
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value is not None:
            a_rows += 1
        if ws.cell(row=r, column=8).value is not None:
            h_rows += 1
    print(f"  Rows with non-empty A: {a_rows}, non-empty H: {h_rows}")
wb.close()
