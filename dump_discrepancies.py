import openpyxl

wb = openpyxl.load_workbook(r"C:\Users\Maxim\Downloads\13-19.07_сопоставлено.xlsx", data_only=True)
with open("discrepancies_dump.txt", "w", encoding="utf-8") as f:
    if "⚠ Несоответствия" in wb.sheetnames:
        ws = wb["⚠ Несоответствия"]
        f.write("--- Discrepancies Sheet ---\n")
        for r in range(1, ws.max_row + 1):
            vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
            if any(v is not None for v in vals):
                f.write(f"Row {r:02d}: " + " | ".join([str(v) if v is not None else "" for v in vals]) + "\n")
wb.close()
print("Saved dump to discrepancies_dump.txt")
