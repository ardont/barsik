# -*- coding: utf-8 -*-
"""
Модуль экспорта результатов сверки в формат Excel
"""

import os
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from models import ServiceItem, ReconciliationSummary
from config import EXCEL_STYLES

def export_to_excel(
    tp_items: list,
    bt_items: list,
    matches: list,
    unmatched_tp: list,
    unmatched_bt: list,
    summary: ReconciliationSummary,
    output_path: str
) -> None:
    """
    Экспортирует результаты сверки в многостраничный Excel файл с цветовым оформлением
    """
    wb = openpyxl.Workbook()
    
    # Удаляем стандартный лист
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
        
    # Создаем листы
    ws_summary = wb.create_sheet("📋 Сводка")
    ws_all = wb.create_sheet("📊 Все сопоставления")
    ws_mismatches = wb.create_sheet("⚠ Несоответствия")
    
    # Стили
    font_family = EXCEL_STYLES["font_name"]
    header_fill = PatternFill(start_color=EXCEL_STYLES["header_fill"], end_color=EXCEL_STYLES["header_fill"], fill_type="solid")
    header_font = Font(name=font_family, size=11, bold=True, color=EXCEL_STYLES["header_font_color"])
    
    matched_fill = PatternFill(start_color=EXCEL_STYLES["matched_fill"], end_color=EXCEL_STYLES["matched_fill"], fill_type="solid")
    unmatched_fill = PatternFill(start_color=EXCEL_STYLES["unmatched_fill"], end_color=EXCEL_STYLES["unmatched_fill"], fill_type="solid")
    discrepancy_fill = PatternFill(start_color=EXCEL_STYLES["discrepancy_fill"], end_color=EXCEL_STYLES["discrepancy_fill"], fill_type="solid")
    
    border_side = Side(border_style="thin", color=EXCEL_STYLES["border_color"])
    thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    # ----------------------------------------------------
    # 1. Заполнение листа «Сводка»
    # ----------------------------------------------------
    ws_summary.views.sheetView[0].showGridLines = True
    
    # Заголовок
    ws_summary.cell(row=2, column=2, value="РЕЗУЛЬТАТЫ СВЕРКИ ДАННЫХ").font = Font(name=font_family, size=16, bold=True, color="1F497D")
    ws_summary.cell(row=3, column=2, value=f"Дата проведения сверки: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}").font = Font(name=font_family, size=10, italic=True)
    
    # Таблица статистики
    stats = [
        ("Показатель", "Количество", "Сумма (руб.)"),
        ("Всего операций TicketProf", summary.total_tp_count, summary.total_tp_sum),
        ("Всего операций Bars Tour", summary.total_bt_count, summary.total_bt_sum),
        ("Успешно сопоставлено (Стоимость услуг)", summary.matched_tp_count, summary.matched_tp_sum),
        ("Успешно сопоставлено (Итого в Барсе)", summary.matched_bt_count, summary.matched_bt_sum),
        ("В Тикете, нет в Барсе", summary.unmatched_tp_count, summary.unmatched_tp_sum),
        ("В Барсе, нет в Тикете", summary.unmatched_bt_count, summary.unmatched_bt_sum),
        ("Прибыль (Итого в Барсе - Стоимость услуг)", "", summary.total_profit)
    ]
    
    start_row = 5
    for i, row_data in enumerate(stats):
        r = start_row + i
        for c, val in enumerate(row_data, 2):
            cell = ws_summary.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if i == 0:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = align_center
            else:
                cell.font = Font(name=font_family, size=10, bold=(i == len(stats)-1))
                if c == 2:
                    cell.alignment = align_left
                else:
                    cell.alignment = align_right
                    if isinstance(val, (int, float)):
                        cell.number_format = '#,##0.00'
                        
    # Настройка размеров колонок на сводке
    ws_summary.column_dimensions['B'].width = 45
    ws_summary.column_dimensions['C'].width = 18
    ws_summary.column_dimensions['D'].width = 22
    
    # ----------------------------------------------------
    # 2. Заполнение листа «Все сопоставления»
    # ----------------------------------------------------
    ws_all.views.sheetView[0].showGridLines = True
    
    headers = [
        "Сквозной ID", "Тип услуги",
        "Документ TicketProf", "Описание TicketProf", "Стоимость услуг",
        "Документ Bars Tour", "Описание Bars Tour", "Итого в Барсе",
        "Прибыль", "Метод сопоставления", "Статус"
    ]
    
    # Пишем заголовки
    for col_idx, h in enumerate(headers, 1):
        cell = ws_all.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = thin_border
        
    # Собираем данные
    row_idx = 2
    
    # Добавляем сначала сопоставления
    for tp, bt, method, score in matches:
        profit = bt.amount - tp.allocated_amount
        status_text = tp.get_status_text(bt)
        
        row_fill = matched_fill
        if status_text in ["Нетипичная маржа", "Несовпадение по суммам"]:
            row_fill = discrepancy_fill
            
        tp_ids = list(tp.ids) if tp.ids else []
        bt_ids = list(bt.ids) if bt.ids else []
        all_ids = sorted(list(set(tp_ids + bt_ids)))
        tp_id = ", ".join(all_ids) if all_ids else "N/A"
        
        row_data = [
            tp_id, tp.service_type,
            tp.doc, tp.desc, tp.allocated_amount,
            bt.doc, bt.desc, bt.amount,
            profit, method, status_text
        ]
        
        for col_idx, val in enumerate(row_data, 1):
            cell = ws_all.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.fill = row_fill
            cell.font = Font(name=font_family, size=9)
            
            # Выравнивание
            if col_idx in [1, 2, 3, 6, 10, 11]:
                cell.alignment = align_center
            elif col_idx in [4, 7]:
                cell.alignment = align_left
            else:
                cell.alignment = align_right
                if isinstance(val, (int, float)):
                    cell.number_format = '#,##0.00'
                    
        row_idx += 1
        
    # Добавляем нераспределенные TP
    for tp in unmatched_tp:
        tp_ids = sorted(list(tp.ids)) if tp.ids else []
        tp_id = ", ".join(tp_ids) if tp_ids else "N/A"
        status_text = tp.get_status_text(None)
        
        row_data = [
            tp_id, tp.service_type,
            tp.doc, tp.desc, tp.allocated_amount,
            "", "Отсутствует в Bars Tour", 0.0,
            0.0, "Не сопоставлено", status_text
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws_all.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.fill = unmatched_fill
            cell.font = Font(name=font_family, size=9)
            
            if col_idx in [1, 2, 3, 6, 10, 11]:
                cell.alignment = align_center
            elif col_idx in [4, 7]:
                cell.alignment = align_left
            else:
                cell.alignment = align_right
                if isinstance(val, (int, float)):
                    cell.number_format = '#,##0.00'
        row_idx += 1
        
    # Добавляем нераспределенные BT
    for bt in unmatched_bt:
        bt_ids = sorted(list(bt.ids)) if bt.ids else []
        bt_id = ", ".join(bt_ids) if bt_ids else "N/A"
        status_text = bt.get_status_text(None)
        
        row_fill = matched_fill if status_text == "Норма (Сбор в БТ)" else unmatched_fill
        
        row_data = [
            bt_id, bt.service_type,
            "", "Отсутствует в TicketProf", 0.0,
            bt.doc, bt.desc, bt.amount,
            0.0, "Не сопоставлено", status_text
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws_all.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.fill = row_fill
            cell.font = Font(name=font_family, size=9)
            
            if col_idx in [1, 2, 3, 6, 10, 11]:
                cell.alignment = align_center
            elif col_idx in [4, 7]:
                cell.alignment = align_left
            else:
                cell.alignment = align_right
                if isinstance(val, (int, float)):
                    cell.number_format = '#,##0.00'
        row_idx += 1

    # ----------------------------------------------------
    # 3. Заполнение листа «Несоответствия»
    # ----------------------------------------------------
    ws_mismatches.views.sheetView[0].showGridLines = True
    
    # Заголовки те же
    for col_idx, h in enumerate(headers, 1):
        cell = ws_mismatches.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = thin_border
        
    m_row_idx = 2
    
    # Копируем красные и желтые строки с основного листа
    for r in range(2, row_idx):
        status_val = ws_all.cell(row=r, column=11).value
        if status_val in ["Нетипичная маржа", "В Тикете, нет в Барсе", "В Барсе, нет в Тикете", "Несовпадение по суммам"]:
            row_fill = discrepancy_fill if status_val in ["Нетипичная маржа", "Несовпадение по суммам"] else unmatched_fill
            for col_idx in range(1, 12):
                cell_src = ws_all.cell(row=r, column=col_idx)
                cell_dst = ws_mismatches.cell(row=m_row_idx, column=col_idx, value=cell_src.value)
                cell_dst.border = thin_border
                cell_dst.fill = row_fill
                cell_dst.font = Font(name=font_family, size=9)
                
                if col_idx in [1, 2, 3, 6, 10, 11]:
                    cell_dst.alignment = align_center
                elif col_idx in [4, 7]:
                    cell_dst.alignment = align_left
                else:
                    cell_dst.alignment = align_right
                    
                if cell_src.number_format:
                    cell_dst.number_format = cell_src.number_format
            m_row_idx += 1
            
    # Автоподбор ширины столбцов
    for ws_cur in [ws_all, ws_mismatches]:
        for col in ws_cur.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws_cur.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)
            
    wb.save(output_path)
