import openpyxl
from config import COL_MAP_SINGLE_BT
from engine.normalizer import clean_text, classify_service, extract_identifiers
from engine.loader import to_float, ServiceItem

file_path = r"C:\Users\Maxim\Downloads\20-26_07_26.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb.active

col_map = COL_MAP_SINGLE_BT
bt_items = []
current_doc = None
current_date = None

for r in range(2, ws.max_row + 1):
    val_date = ws.cell(row=r, column=col_map["date"]).value
    val_doc = ws.cell(row=r, column=col_map["doc"]).value
    val_amt = ws.cell(row=r, column=col_map["amount"]).value
    val_profit = ws.cell(row=r, column=col_map["profit"]).value
    val_net = ws.cell(row=r, column=col_map["net"]).value
    
    is_header = False
    if val_date is not None and val_doc is not None:
        if any(kw in str(val_doc) for kw in ["Приход", "Оплата", "Возврат", "Принято"]):
            is_header = True
            
    if is_header:
        current_date = str(val_date).strip()
        current_doc = str(val_doc).strip()
    elif val_date is not None and val_doc is None:
        desc = str(val_date).strip()
        if not any(kw in desc for kw in ["Обороты за период", "Сальдо конечное", "Итого"]):
            j_val = to_float(val_amt)
            amt = j_val if j_val is not None else 0.0
            if amt == 0.0 and not extract_identifiers(desc):
                continue
                
            profit_val = to_float(val_profit)
            net_val = to_float(val_net)
            
            bt_items.append(ServiceItem(
                row=r,
                date=current_date or "",
                doc=current_doc or "",
                desc=desc,
                clean_desc=clean_text(desc),
                service_type=classify_service(desc),
                amount=amt,
                allocated_amount=amt,
                ids=extract_identifiers(desc),
                profit=profit_val,
                net=net_val,
                source="BT"
            ))

print("Raw un-grouped bt_items for 555-2397261625:")
for x in bt_items:
    if "555-2397261625" in str(x.ids):
        print(f"  Row {x.row:2d} | Amt: {x.amount} | Doc: {repr(x.doc)} | Desc: {repr(x.desc[:30])}")

from collections import defaultdict
grouped_bt = defaultdict(list)
for item in bt_items:
    key = (item.doc, item.clean_desc, item.service_type, frozenset(item.ids))
    grouped_bt[key].append(item)

print("\nGrouped results for 555-2397261625:")
for key, items in grouped_bt.items():
    if "555-2397261625" in str(key):
        print(f"Key: {key}")
        print(f"Items count: {len(items)}")
        for x in items:
            print(f"  Row {x.row:2d} | Amt: {x.amount}")
