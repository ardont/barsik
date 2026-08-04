import openpyxl
from config import COL_MAP_SINGLE_BT

file_path = r"C:\Users\Maxim\Downloads\20-26_07_26.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb.active

col_map = COL_MAP_SINGLE_BT
current_doc = None
current_date = None

for r in range(2, ws.max_row + 1):
    val_date = ws.cell(row=r, column=col_map["date"]).value
    val_doc = ws.cell(row=r, column=col_map["doc"]).value
    val_amt = ws.cell(row=r, column=col_map["amount"]).value
    
    is_header = False
    if val_date is not None and val_doc is not None:
        if any(kw in str(val_doc) for kw in ["Приход", "Оплата", "Возврат", "Принято"]):
            is_header = True
            
    if is_header:
        current_date = str(val_date).strip()
        current_doc = str(val_doc).strip()
        if 60 <= r <= 75:
            print(f"HEADER at row {r}: doc={current_doc}")
    elif val_date is not None and val_doc is None:
        desc = str(val_date).strip()
        if not any(kw in desc for kw in ["Обороты за период", "Сальдо конечное", "Итого"]):
            if 60 <= r <= 75:
                print(f"ITEM at row {r}: amt={val_amt} | doc={current_doc} | desc={repr(desc[:30])}")
    else:
        if 60 <= r <= 75:
            print(f"OTHER at row {r}: val_date={repr(val_date)} | val_doc={repr(val_doc)}")
