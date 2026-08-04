import openpyxl
import re
from collections import defaultdict
from typing import List, Tuple, Any, Optional, Set
from dataclasses import dataclass, field

# Let's import config and normalizer
from config import COL_MAP_SINGLE_TP, COL_MAP_SINGLE_BT
from engine.normalizer import clean_text, classify_service, extract_identifiers
from models import ServiceItem, ReconciliationSummary
from engine.loader import to_float

def fixed_parse_tp_sheet(ws: Any, col_map: dict) -> Tuple[List[ServiceItem], dict]:
    tp_items = []
    doc_sums = {}
    current_doc = None
    current_date = None
    
    for r in range(2, ws.max_row + 1):
        val_date = ws.cell(row=r, column=col_map["date"]).value
        val_doc = ws.cell(row=r, column=col_map["doc"]).value
        val_debit = ws.cell(row=r, column=col_map["debit"]).value
        val_credit = ws.cell(row=r, column=col_map["credit"]).value
        
        is_header = False
        if val_date is not None and val_doc is not None:
            v_doc_str = str(val_doc).strip()
            if any(kw in v_doc_str for kw in ["Продажа", "Возврат", "Корректировка"]):
                # Исключаем платежи и сальдо из шапок документов
                if not any(skip in v_doc_str for skip in ["Оплата", "Сальдо", "Банк", "Выписка"]):
                    is_header = True
                
        if is_header:
            current_date = str(val_date).strip()
            current_doc = str(val_doc).strip()
            c_val = to_float(val_debit)
            d_val = to_float(val_credit)
            current_doc_amt = c_val if c_val is not None else (-d_val if d_val is not None else 0.0)
            doc_sums[(current_date, current_doc)] = current_doc_amt
        elif val_date is not None and val_doc is None:
            desc = str(val_date).strip()
            # Фильтруем служебные строки, сальдо и оплаты
            if any(kw in desc for kw in ["Обороты за период", "Сальдо конечное", "Сальдо начальное", "Сальдо", "Итого", "Оплата", "Оплата по счету"]):
                continue
                
            c_val = to_float(val_debit)
            d_val = to_float(val_credit)
            amt = c_val if c_val is not None else (-d_val if d_val is not None else 0.0)
            
            tp_items.append(ServiceItem(
                row=r,
                date=current_date or "",
                doc=current_doc or "",
                desc=desc,
                clean_desc=clean_text(desc),
                service_type=classify_service(desc),
                amount=amt,
                allocated_amount=amt,
                ids=extract_identifiers(desc),
                source="TP"
            ))
            
    return tp_items, doc_sums

def fixed_parse_bt_sheet(ws: Any, col_map: dict) -> List[ServiceItem]:
    bt_items = []
    current_doc = None
    current_date = None
    
    for r in range(2, ws.max_row + 1):
        val_date = ws.cell(row=r, column=col_map["date"]).value
        val_doc = ws.cell(row=r, column=col_map["doc"]).value
        val_amt = ws.cell(row=r, column=col_map["amount"]).value
        val_profit = ws.cell(row=r, column=col_map["profit"]).value
        val_net = ws.cell(row=r, column=col_map["net"]).value
        
        is_header = False
        if val_date is not None and val_doc is not None:
            v_doc_str = str(val_doc).strip()
            if any(kw in v_doc_str for kw in ["Приход", "Возврат", "Принято"]):
                if not any(skip in v_doc_str for skip in ["Оплата", "Сальдо", "Банк", "Выписка"]):
                    is_header = True
                
        if is_header:
            current_date = str(val_date).strip()
            current_doc = str(val_doc).strip()
        elif val_date is not None and val_doc is None:
            desc = str(val_date).strip()
            if any(kw in desc for kw in ["Обороты за период", "Сальдо конечное", "Сальдо начальное", "Сальдо", "Итого", "Оплата", "Оплата по счету"]):
                continue
                
            j_val = to_float(val_amt)
            amt = j_val if j_val is not None else 0.0
            if amt == 0.0 and not extract_identifiers(desc):
                continue
                
            profit_val = to_float(val_profit)
            net_val = to_float(val_net)
            
            bt_items.append(ServiceItem(
                row=r,
                date=current_date or "",
                doc=current_doc or "",
                desc=desc,
                clean_desc=clean_text(desc),
                service_type=classify_service(desc),
                amount=amt,
                allocated_amount=amt,
                ids=extract_identifiers(desc),
                profit=profit_val,
                net=net_val,
                source="BT"
            ))
            
    # Группировка элементов Bars Tour
    grouped_bt = defaultdict(list)
    for item in bt_items:
        key = (item.doc, item.clean_desc, item.service_type, frozenset(item.ids))
        grouped_bt[key].append(item)
        
    final_bt_items = []
    for key, items in grouped_bt.items():
        if len(items) == 1:
            final_bt_items.append(items[0])
        else:
            sorted_items = sorted(items, key=lambda x: abs(x.amount))
            best_item = sorted_items[-1]
            
            merged_ids = set()
            for x in items:
                merged_ids.update(x.ids)
            best_item.ids = merged_ids
            
            # Проверяем уникальные комбинации сумм
            amounts = [x.amount for x in items]
            
            if len(items) == 2:
                # 2 элемента (например 14462 + 30)
                best_item.amount = sum(amounts)
                best_item.allocated_amount = best_item.amount
            else:
                # Если элементов >= 3 (например 14462, 30, 14462_dup)
                # Вычисляем сумму подстрок (уникальных компонент)
                unique_positive_amounts = list(set(amounts))
                if len(unique_positive_amounts) > 1:
                    best_item.amount = sum(unique_positive_amounts)
                else:
                    best_item.amount = sorted_items[-1].amount
                best_item.allocated_amount = best_item.amount
                
            best_item.row = items[0].row
            final_bt_items.append(best_item)
            
    return final_bt_items

# Test running fixed functions on 20-26_07_26.xlsx
wb = openpyxl.load_workbook(r"C:\Users\Maxim\Downloads\20-26_07_26.xlsx", data_only=True)
ws = wb.active

tp_items, _ = fixed_parse_tp_sheet(ws, COL_MAP_SINGLE_TP)
bt_items = fixed_parse_bt_sheet(ws, COL_MAP_SINGLE_BT)

print(f"Fixed TP items count: {len(tp_items)}")
print(f"Fixed BT items count: {len(bt_items)}")

# Check ticket 555-2397261625 in both
tp_1625 = [x for x in tp_items if "555-2397261625" in x.ids and x.service_type == "Flight"]
bt_1625 = [x for x in bt_items if "555-2397261625" in x.ids and x.service_type == "Flight"]

print("\nTP 1625:", [(x.row, x.amount, x.desc[:30]) for x in tp_1625])
print("BT 1625:", [(x.row, x.amount, x.desc[:30]) for x in bt_1625])

# Check hotels
tp_hotels = [x for x in tp_items if x.service_type == "Hotel"]
bt_hotels = [x for x in bt_items if x.service_type == "Hotel"]

print("\nTP Hotels:", [(x.row, x.amount, x.desc[:30]) for x in tp_hotels])
print("BT Hotels:", [(x.row, x.amount, x.desc[:30]) for x in bt_hotels])
