import openpyxl
import re

file_path = r"C:\Users\Maxim\Downloads\08.07_1.xlsx"

# (Copied functions from test_matcher_intersection.py)
def classify_service(text):
    t = text.lower()
    if any(kw in t for kw in ["проживание", "отель", "г-ца", "гостиница"]):
        return "Hotel"
    if any(kw in t for kw in ["выбор места", "место", "mco"]):
        return "Seat"
    if any(kw in t for kw in ["услуга ито", "сервисный сбор", "ито"]):
        return "Fee"
    if any(kw in t for kw in ["штраф", "удержания", "удержание"]):
        return "Penalty"
    if any(kw in t for kw in ["авиабилет", "билет", "возврат авиабилета"]):
        return "Flight"
    return "Other"

def clean_text(text):
    if not text:
        return ""
    t = text.lower()
    t = t.replace("г-ца", "гостиница").replace("г.", "город ").replace("улица", "ул")
    t = t.replace("авиабилет", "билет")
    t = re.sub(r'\b\d{2}\.\d{2}\.(?:\d{4}|\d{2})\b', '', t)
    t = re.sub(r'\bс\s+\d{2}\.\d{2}\s+по\s+\d{2}\.\d{2}\b', '', t)
    t = re.sub(r'\([a-z\s\/]+:?\s*(?:№)?\s*\d*(?:-\d*)?\)', '', t)
    t = re.sub(r'\([a-z0-9\s\/\-\:\.\,]+\)', '', t)
    t = re.sub(r'\b[a-z]+/[a-z]+\b', '', t)
    t = re.sub(r'\b\d{3}-\d{10}\b', '', t)
    t = re.sub(r'\b(?:заказ|№)?\s*\d+\b', '', t)
    t = re.sub(r'[^\w\s]', ' ', t)
    t = re.sub(r'\s+', ' ', t).strip()
    return t

def extract_identifiers(text):
    if not text:
        return set()
    tickets = re.findall(r'\b\d{3}-\d{10}\b', text)
    orders = re.findall(r'(?:Заказ|заказ|заказа|№)\s*(\d{7,10})\b', text)
    return set(tickets + orders)

wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb["Лист2"]

tp_items = []
current_doc_tp = None
current_date_tp = None
for r in range(2, ws.max_row + 1):
    val_a = ws.cell(row=r, column=1).value
    val_b = ws.cell(row=r, column=2).value
    val_c = ws.cell(row=r, column=3).value
    val_d = ws.cell(row=r, column=4).value
    
    is_header = False
    if val_a is not None and val_b is not None:
        if any(kw in str(val_b) for kw in ["Продажа", "Оплата", "Возврат", "Корректировка"]):
            is_header = True
            
    if is_header:
        current_date_tp = val_a
        current_doc_tp = val_b
    elif val_a is not None and val_b is None:
        desc = str(val_a).strip()
        if any(kw in desc for kw in ["Обороты за период", "Сальдо конечное", "Итого"]):
            continue
        amt = val_c if val_c is not None else (-val_d if val_d is not None else 0.0)
        tp_items.append({
            "row": r,
            "date": current_date_tp,
            "doc": current_doc_tp,
            "desc": desc,
            "clean_desc": clean_text(desc),
            "service_type": classify_service(desc),
            "amount": amt,
            "ids": extract_identifiers(desc),
            "matched": False
        })

bt_items = []
current_doc_bt = None
current_date_bt = None
for r in range(2, ws.max_row + 1):
    val_h = ws.cell(row=r, column=8).value
    val_i = ws.cell(row=r, column=9).value
    val_j = ws.cell(row=r, column=10).value
    val_l = ws.cell(row=r, column=12).value
    val_m = ws.cell(row=r, column=13).value
    
    is_header = False
    if val_h is not None and val_i is not None:
        if any(kw in str(val_i) for kw in ["Приход", "Оплата", "Возврат", "Принято"]):
            is_header = True
            
    if is_header:
        current_date_bt = val_h
        current_doc_bt = val_i
    elif val_h is not None and val_i is None:
        desc = str(val_h).strip()
        if any(kw in desc for kw in ["Обороты за период", "Сальдо конечное", "Итого"]):
            continue
        amt = val_j if val_j is not None else 0.0
        if amt == 0.0 and not extract_identifiers(desc):
            continue
        bt_items.append({
            "row": r,
            "date": current_date_bt,
            "doc": current_doc_bt,
            "desc": desc,
            "clean_desc": clean_text(desc),
            "service_type": classify_service(desc),
            "amount": amt,
            "profit": val_l,
            "net": val_m,
            "ids": extract_identifiers(desc),
            "matched": False
        })

# Do matching
for tp in tp_items:
    if tp["ids"]:
        bt_matches = [b for b in bt_items if b["ids"].intersection(tp["ids"]) and b["service_type"] == tp["service_type"] and not b["matched"]]
        if bt_matches:
            best_bt = min(bt_matches, key=lambda b: abs(b["amount"] - tp["amount"]))
            tp["matched"] = True
            best_bt["matched"] = True

for tp in tp_items:
    if not tp["matched"]:
        bt_matches = [b for b in bt_items if b["clean_desc"] == tp["clean_desc"] and b["service_type"] == tp["service_type"] and not b["matched"]]
        if bt_matches:
            best_bt = min(bt_matches, key=lambda b: abs(b["amount"] - tp["amount"]))
            tp["matched"] = True
            best_bt["matched"] = True

import difflib
for tp in tp_items:
    if not tp["matched"]:
        unmatched_bt = [b for b in bt_items if not b["matched"] and b["service_type"] == tp["service_type"]]
        best_bt = None
        best_score = 0.0
        for bt in unmatched_bt:
            score = difflib.SequenceMatcher(None, tp["clean_desc"], bt["clean_desc"]).ratio()
            if score > best_score:
                best_score = score
                best_bt = bt
        if best_bt and best_score >= 0.75:
            tp["matched"] = True
            best_bt["matched"] = True

print("=== UNMATCHED TICKETPROF ITEMS ===")
for tp in tp_items:
    if not tp["matched"]:
        print(f"Row {tp['row']} | {tp['service_type']} | Amt: {tp['amount']} | IDs: {tp['ids']} | {tp['desc'][:80]}")

print("\n=== UNMATCHED BARS TOUR ITEMS ===")
for bt in bt_items:
    if not bt["matched"]:
        print(f"Row {bt['row']} | {bt['service_type']} | Amt: {bt['amount']} | IDs: {bt['ids']} | {bt['desc'][:80]}")

wb.close()
