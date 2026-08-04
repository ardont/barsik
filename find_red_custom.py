import openpyxl

wb = openpyxl.load_workbook(r"C:\Users\Maxim\Downloads\20-26_07_26.xlsx", data_only=False)
ws = wb.active

for r in range(1, ws.max_row + 1):
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(r, c)
        
        fill_rgb = None
        if cell.fill and cell.fill.fgColor:
            fill_rgb = str(cell.fill.fgColor.rgb or cell.fill.fgColor.theme or cell.fill.fgColor.indexed)
            
        font_rgb = None
        if cell.font and cell.font.color:
            font_rgb = str(cell.font.color.rgb or cell.font.color.theme or cell.font.color.indexed)
            
        comment = cell.comment.text if cell.comment else ""
        
        # Filter out 00000000 / None / standard cell styles
        is_fill_custom = fill_rgb and fill_rgb not in ['00000000', 'None', 'FFFFFFFF', '00FFFFFF']
        is_font_custom = font_rgb and font_rgb not in ['00000000', 'None', 'FF000000', '000000', 'FFFFFFFF']
        
        if is_fill_custom or is_font_custom or comment:
            print(f"Row {r:3d}, Col {c:2d} ({cell.coordinate}): val={repr(cell.value)} | fill={fill_rgb} | font={font_rgb} | comment={repr(comment)}")
