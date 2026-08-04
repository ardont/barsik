import openpyxl
from config import COL_MAP_SINGLE_BT

file_path = r"C:\Users\Maxim\Downloads\20-26_07_26.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb.active

for r in [68, 69, 70, 71]:
    v_date = ws.cell(row=r, column=COL_MAP_SINGLE_BT["date"]).value
    v_doc = ws.cell(row=r, column=COL_MAP_SINGLE_BT["doc"]).value
    v_amt = ws.cell(row=r, column=COL_MAP_SINGLE_BT["amount"]).value
    print(f"Row {r}: v_date={repr(v_date)} | v_doc={repr(v_doc)} | v_amt={repr(v_amt)}")
