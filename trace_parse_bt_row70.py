import openpyxl
from config import COL_MAP_SINGLE_BT
from engine.normalizer import clean_text, classify_service, extract_identifiers
from engine.loader import to_float

file_path = r"C:\Users\Maxim\Downloads\20-26_07_26.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb.active

col_map = COL_MAP_SINGLE_BT

for r in range(65, 75):
    val_date = ws.cell(row=r, column=col_map["date"]).value
    val_doc = ws.cell(row=r, column=col_map["doc"]).value
    val_amt = ws.cell(row=r, column=col_map["amount"]).value
    val_profit = ws.cell(row=r, column=col_map["profit"]).value
    val_net = ws.cell(row=r, column=col_map["net"]).value
    
    is_header = False
    if val_date is not None and val_doc is not None:
        if any(kw in str(val_doc) for kw in ["Приход", "Оплата", "Возврат", "Принято"]):
            is_header = True
            
    print(f"Row {r:2d}: val_date={repr(val_date)[:30]} | val_doc={repr(val_doc)} | is_header={is_header}")
    if not is_header and val_date is not None and val_doc is None:
        desc = str(val_date).strip()
        skipping = any(kw in desc for kw in ["Обороты за период", "Сальдо конечное", "Итого"])
        j_val = to_float(val_amt)
        amt = j_val if j_val is not None else 0.0
        ids = extract_identifiers(desc)
        svc_type = classify_service(desc)
        print(f"        desc_clean={repr(desc[:30])} | amt={amt} | skipping={skipping} | ids={ids} | type={svc_type}")
