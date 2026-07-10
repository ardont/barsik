import openpyxl

file_path = r"C:\Users\Maxim\Downloads\08.07_1.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb["Лист2"]

print("=== CHECKING ALL ROWS WITH VALUES IN L OR M ===")
print("Row | Col H (Description) | Col J (Amount) | Col L | Col M")
print("-" * 100)
for r in range(2, ws.max_row + 1):
    val_l = ws.cell(row=r, column=12).value
    val_m = ws.cell(row=r, column=13).value
    if val_l is not None or val_m is not None:
        desc = ws.cell(row=r, column=8).value
        amt = ws.cell(row=r, column=10).value
        # truncate description
        desc_trunc = str(desc)[:40] if desc else ""
        str_l = str(val_l) if val_l is not None else ""
        str_m = str(val_m) if val_m is not None else ""
        str_amt = str(amt) if amt is not None else ""
        print(f"{r:3d} | {desc_trunc:<40} | {str_amt:<14} | {str_l:<8} | {str_m:<8}")
wb.close()
