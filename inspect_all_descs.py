import openpyxl
import re

wb = openpyxl.load_workbook(r'C:\Users\Maxim\Downloads\13-19.07.xlsx', data_only=True)
ws = wb.active

print("=== ALL DESCRIPTIONS IN 13-19.07.xlsx ===")
for r in range(1, ws.max_row + 1):
    tp_desc = ws.cell(r, 1).value
    bt_desc = ws.cell(r, 8).value
    
    if tp_desc:
        print(f"Row {r:2d} TP: {tp_desc}")
    if bt_desc:
        print(f"Row {r:2d} BT: {bt_desc}")
