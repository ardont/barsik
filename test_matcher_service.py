import openpyxl
import re
from collections import defaultdict
import difflib

file_path = r"C:\Users\Maxim\Downloads\08.07_1.xlsx"

def classify_service(text):
    t = text.lower()
    if any(kw in t for kw in ["проживание", "отель", "г-ца", "гостиница"]):
        return "Hotel"
    if any(kw in t for kw in ["выбор места", "место", "mco"]):
        return "Seat"
    if any(kw in t for kw in ["услуга ито", "сервисный сбор", "ито"]):
        return "Fee"
    if any(kw in t for kw in ["штраф", "удержания", "удержание"]):
        return "Penalty"
    if any(kw in t for kw in ["авиабилет", "билет", "возврат авиабилета"]):
        return "Flight"
    return "Other"

def clean_text(text):
    if not text:
        return ""
    t = text.lower()
    t = t.replace("г-ца", "гостиница").replace("г.", "город ").replace("улица", "ул")
    t = t.replace("авиабилет", "билет")
    t = re.sub(r'\b\d{2}\.\d{2}\.(?:\d{4}|\d{2})\b', '', t)
    t = re.sub(r'\bс\s+\d{2}\.\d{2}\s+по\s+\d{2}\.\d{2}\b', '', t)
    t = re.sub(r'\([a-z\s\/]+:?\s*(?:№)?\s*\d*(?:-\d*)?\)', '', t)
    t = re.sub(r'\([a-z0-9\s\/\-\:\.\,]+\)', '', t)
    t = re.sub(r'\b[a-z]+/[a-z]+\b', '', t)
    t = re.sub(r'\b\d{3}-\d{10}\b', '', t)
    t = re.sub(r'\b(?:заказ|№)?\s*\d+\b', '', t)
    t = re.sub(r'[^\w\s]', ' ', t)
    t = re.sub(r'\s+', ' ', t).strip()
    return t

def extract_id(text):
    if not text:
        return None, None
    ticket_match = re.search(r'\b\d{3}-\d{10}\b', text)
    if ticket_match:
        return "ticket", ticket_match.group(0)
    order_match = re.search(r'(?:Заказ|заказ)\s*(?:№)?\s*(\d+)', text)
    if order_match:
        return "order", order_match.group(1)
    return None, None

wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb["Лист2"]

# Parse TicketProf (Columns A-C)
tp_items = []
current_doc_tp = None
current_date_tp = None
for r in range(2, ws.max_row + 1):
    val_a = ws.cell(row=r, column=1).value
    val_b = ws.cell(row=r, column=2).value
    val_c = ws.cell(row=r, column=3).value
    val_d = ws.cell(row=r, column=4).value
    
    is_header = False
    if val_a is not None and val_b is not None:
        if any(kw in str(val_b) for kw in ["Продажа", "Оплата", "Возврат", "Корректировка"]):
            is_header = True
            
    if is_header:
        current_date_tp = val_a
        current_doc_tp = val_b
    elif val_a is not None and val_b is None:
        desc = str(val_a).strip()
        if any(kw in desc for kw in ["Обороты за период", "Сальдо конечное", "Итого"]):
            continue
        id_type, id_val = extract_id(desc)
        amt = val_c if val_c is not None else (-val_d if val_d is not None else 0.0)
        tp_items.append({
            "row": r,
            "date": current_date_tp,
            "doc": current_doc_tp,
            "desc": desc,
            "clean_desc": clean_text(desc),
            "service_type": classify_service(desc),
            "amount": amt,
            "id_type": id_type,
            "id_val": id_val,
            "matched": False
        })

# Parse Bars Tour (Columns H-M)
bt_items = []
current_doc_bt = None
current_date_bt = None
for r in range(2, ws.max_row + 1):
    val_h = ws.cell(row=r, column=8).value
    val_i = ws.cell(row=r, column=9).value
    val_j = ws.cell(row=r, column=10).value
    val_l = ws.cell(row=r, column=12).value
    val_m = ws.cell(row=r, column=13).value
    
    is_header = False
    if val_h is not None and val_i is not None:
        if any(kw in str(val_i) for kw in ["Приход", "Оплата", "Возврат", "Принято"]):
            is_header = True
            
    if is_header:
        current_date_bt = val_h
        current_doc_bt = val_i
    elif val_h is not None and val_i is None:
        desc = str(val_h).strip()
        if any(kw in desc for kw in ["Обороты за период", "Сальдо конечное", "Итого"]):
            continue
        id_type, id_val = extract_id(desc)
        amt = val_j if val_j is not None else 0.0
        if amt == 0.0 and not id_val:
            continue
        bt_items.append({
            "row": r,
            "date": current_date_bt,
            "doc": current_doc_bt,
            "desc": desc,
            "clean_desc": clean_text(desc),
            "service_type": classify_service(desc),
            "amount": amt,
            "profit": val_l,
            "net": val_m,
            "id_type": id_type,
            "id_val": id_val,
            "matched": False
        })

# Matching Engine Simulation with Service Type filter
matches = []

# Step 1: Match by ID and Service Type
for tp in tp_items:
    if tp["id_val"]:
        bt_matches = [b for b in bt_items if b["id_val"] == tp["id_val"] and b["service_type"] == tp["service_type"] and not b["matched"]]
        if bt_matches:
            best_bt = min(bt_matches, key=lambda b: abs(b["amount"] - tp["amount"]))
            tp["matched"] = True
            best_bt["matched"] = True
            matches.append((tp, best_bt, "ID + Service Type Match", 1.0))

# Step 2: Match by exact clean description
for tp in tp_items:
    if not tp["matched"]:
        bt_matches = [b for b in bt_items if b["clean_desc"] == tp["clean_desc"] and b["service_type"] == tp["service_type"] and not b["matched"]]
        if bt_matches:
            best_bt = min(bt_matches, key=lambda b: abs(b["amount"] - tp["amount"]))
            tp["matched"] = True
            best_bt["matched"] = True
            matches.append((tp, best_bt, "Exact Clean Name Match", 1.0))

# Step 3: Fuzzy matching with same service type
for tp in tp_items:
    if not tp["matched"]:
        unmatched_bt = [b for b in bt_items if not b["matched"] and b["service_type"] == tp["service_type"]]
        best_bt = None
        best_score = 0.0
        for bt in unmatched_bt:
            score = difflib.SequenceMatcher(None, tp["clean_desc"], bt["clean_desc"]).ratio()
            if score > best_score:
                best_score = score
                best_bt = bt
        if best_bt and best_score >= 0.75:
            tp["matched"] = True
            best_bt["matched"] = True
            matches.append((tp, best_bt, f"Fuzzy Name Match ({tp['service_type']})", best_score))

matched_tp = len([t for t in tp_items if t["matched"]])
matched_bt = len([b for b in bt_items if b["matched"]])
print(f"Matched TicketProf: {matched_tp}/{len(tp_items)} ({matched_tp/len(tp_items)*100:.1f}%)")
print(f"Matched Bars Tour: {matched_bt}/{len(bt_items)} ({matched_bt/len(bt_items)*100:.1f}%)")

print("\n--- DETAILED EVALUATION OF SPECIFIC ID 555-2393969968 ---")
for tp in tp_items:
    if tp["id_val"] == "555-2393969968":
        print(f"TP Row {tp['row']} ({tp['service_type']}): Clean: {tp['clean_desc']} | Amt: {tp['amount']} | Matched: {tp['matched']}")
for bt in bt_items:
    if bt["id_val"] == "555-2393969968":
        print(f"BT Row {bt['row']} ({bt['service_type']}): Clean: {bt['clean_desc']} | Amt: {bt['amount']} | Matched: {bt['matched']}")

wb.close()
