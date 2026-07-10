import openpyxl

file_path = r"C:\Users\Maxim\Downloads\08.07_1.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb["Лист2"]

print("=== CELL BY CELL DUMP FOR ROWS 14 TO 22 ===")
for r in range(14, 23):
    print(f"\nRow {r}:")
    for c in range(1, 15):
        val = ws.cell(row=r, column=c).value
        if val is not None:
            col_letter = openpyxl.utils.get_column_letter(c)
            print(f"  Col {col_letter} ({c}): {repr(val)}")
wb.close()
