import openpyxl

wb = openpyxl.load_workbook(r"C:\Users\Maxim\Downloads\20-26_07_26.xlsx", data_only=False)
ws = wb.active

out = []
out.append(f"Sheet name: {ws.title}, Rows: {ws.max_row}, Cols: {ws.max_column}")

for r in range(1, ws.max_row + 1):
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(r, c)
        
        fill_rgb = None
        if cell.fill and cell.fill.fgColor:
            fill_rgb = cell.fill.fgColor.rgb or cell.fill.fgColor.theme or cell.fill.fgColor.indexed
            
        font_rgb = None
        if cell.font and cell.font.color:
            font_rgb = cell.font.color.rgb or cell.font.color.theme or cell.font.color.indexed
            
        comment = cell.comment.text if cell.comment else None
        
        s_fill = str(fill_rgb) if fill_rgb is not None else ""
        s_font = str(font_rgb) if font_rgb is not None else ""
        
        if s_fill or s_font or comment or (cell.value and str(cell.value).startswith('=')):
            out.append(f"Row {r:3d}, Col {c:2d} ({cell.coordinate}): val={repr(cell.value)} | fill={s_fill} | font={s_font} | comment={repr(comment)}")

with open("red_report_utf8.txt", "w", encoding="utf-8") as f:
    f.write("\n".join(out))

print("Wrote red_report_utf8.txt")
