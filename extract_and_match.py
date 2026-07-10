import openpyxl
import re
from collections import defaultdict

file_path = r"C:\Users\Maxim\Downloads\08.07_1.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb["Лист2"]

def clean_text(text):
    if not text:
        return ""
    return re.sub(r'\s+', ' ', str(text)).strip()

def extract_id(text):
    if not text:
        return None, None
    # Check for ticket number (e.g., 555-2394406838 or 425-6111246241)
    ticket_match = re.search(r'\b\d{3}-\d{10}\b', text)
    if ticket_match:
        return "ticket", ticket_match.group(0)
    # Check for hotel order number (e.g., Заказ 806133056)
    order_match = re.search(r'(?:Заказ|заказ)\s*(?:№)?\s*(\d+)', text)
    if order_match:
        return "order", order_match.group(1)
    return None, None

ticket_prof_items = []
bars_tour_items = []

# Parse TicketProf (Columns A-C)
current_doc_tp = None
current_date_tp = None
for r in range(2, ws.max_row + 1):
    val_a = ws.cell(row=r, column=1).value
    val_b = ws.cell(row=r, column=2).value
    val_c = ws.cell(row=r, column=3).value
    
    if val_a is not None and val_b is not None:
        # Document header
        current_date_tp = val_a
        current_doc_tp = val_b
    elif val_a is not None and val_b is None and val_c is not None:
        # Description item
        desc = clean_text(val_a)
        id_type, id_val = extract_id(desc)
        ticket_prof_items.append({
            "row": r,
            "date": current_date_tp,
            "doc": current_doc_tp,
            "desc": desc,
            "amount": val_c,
            "id_type": id_type,
            "id_val": id_val
        })

# Parse Bars Tour (Columns H-M)
current_doc_bt = None
current_date_bt = None
for r in range(2, ws.max_row + 1):
    val_h = ws.cell(row=r, column=8).value
    val_i = ws.cell(row=r, column=9).value
    val_j = ws.cell(row=r, column=10).value
    val_l = ws.cell(row=r, column=12).value
    val_m = ws.cell(row=r, column=13).value
    
    if val_h is not None and val_i is not None:
        # Document header
        current_date_bt = val_h
        current_doc_bt = val_i
    elif val_h is not None and val_i is None and val_j is not None:
        # Description item
        desc = clean_text(val_h)
        id_type, id_val = extract_id(desc)
        bars_tour_items.append({
            "row": r,
            "date": current_date_bt,
            "doc": current_doc_bt,
            "desc": desc,
            "amount": val_j,
            "profit": val_l,
            "net": val_m,
            "id_type": id_type,
            "id_val": id_val
        })

print(f"Extracted {len(ticket_prof_items)} items from TicketProf")
print(f"Extracted {len(bars_tour_items)} items from Bars Tour")

# Group by ID (ticket or order)
tp_grouped = defaultdict(list)
bt_grouped = defaultdict(list)

for item in ticket_prof_items:
    if item["id_val"]:
        tp_grouped[item["id_val"]].append(item)

for item in bars_tour_items:
    if item["id_val"]:
        bt_grouped[item["id_val"]].append(item)

all_ids = set(tp_grouped.keys()).union(set(bt_grouped.keys()))
print(f"\nUnique IDs found: {len(all_ids)}")

print("\n--- SAMPLE ID MATCHING ---")
sample_count = 0
for idx in sorted(all_ids):
    if sample_count >= 10:
        break
    tp_list = tp_grouped[idx]
    bt_list = bt_grouped[idx]
    if tp_list and bt_list:
        print(f"\nID: {idx} (Type: {tp_list[0]['id_type'] if tp_list else bt_list[0]['id_type']})")
        print("  TicketProf items:")
        for item in tp_list:
            print(f"    Row {item['row']}: {item['desc'][:60]}... Amount: {item['amount']}")
        print("  Bars Tour items:")
        for item in bt_list:
            print(f"    Row {item['row']}: {item['desc'][:60]}... Amount: {item['amount']}, Profit: {item['profit']}, Net: {item['net']}")
        sample_count += 1

wb.close()
