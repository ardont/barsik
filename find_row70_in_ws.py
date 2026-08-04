import openpyxl
from engine.loader import parse_bt_sheet
from config import COL_MAP_SINGLE_BT

file_path = r"C:\Users\Maxim\Downloads\20-26_07_26.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb.active

bt_items = parse_bt_sheet(ws, COL_MAP_SINGLE_BT)

print("Rows in bt_items:")
for x in bt_items:
    if 65 <= x.row <= 75:
        print(f"Row {x.row:2d} | Amt: {x.amount} | Desc: {repr(x.desc)}")

print("\nDirect openpyxl rows 65-75:")
for r in range(65, 75):
    v_date = ws.cell(row=r, column=8).value
    v_doc = ws.cell(row=r, column=9).value
    v_amt = ws.cell(row=r, column=10).value
    print(f"Excel Row {r:2d} | v_date: {repr(v_date)} | v_doc: {repr(v_doc)} | v_amt: {repr(v_amt)}")
