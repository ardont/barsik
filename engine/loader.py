# -*- coding: utf-8 -*-
"""
Модуль загрузки и первичного парсинга данных из Excel
"""

import openpyxl
import re
from collections import defaultdict
from typing import List, Tuple, Any, Optional
from models import ServiceItem
from engine.normalizer import clean_text, classify_service, extract_identifiers
from config import (
    COL_MAP_SINGLE_TP, COL_MAP_SINGLE_BT,
    COL_MAP_DOUBLE_TP, COL_MAP_DOUBLE_BT
)

def to_float(val: Any) -> Optional[float]:
    """
    Безопасное приведение значения к float с поддержкой запятых, пробелов и текстового мусора
    """
    if val is None or val == "":
        return None
    try:
        s = str(val).replace(" ", "").replace("\xa0", "").replace(",", ".").strip()
        # Извлекаем первое число, если строка начинается с цифры или знака
        match = re.match(r'^[-+]?\d*\.?\d+', s)
        if match:
            return float(match.group(0))
        return float(s)
    except (ValueError, TypeError):
        return None

def load_data(tp_path: str, bt_path: Optional[str] = None) -> Tuple[List[ServiceItem], List[ServiceItem]]:
    """
    Загружает файлы Excel и разбирает их на два потока: TicketProf и Bars Tour.
    Поддерживает:
    - Сводный режим (bt_path не передан): парсинг TP из A-D, BT из H-M в файле tp_path.
    - Раздельный режим (переданы оба пути): парсинг TP из tp_path (A-D), BT из bt_path (A-F).
    """
    
    tp_items: List[ServiceItem] = []
    bt_items: List[ServiceItem] = []
    doc_sums = {}
    
    # Режим сводного файла
    if not bt_path:
        wb = openpyxl.load_workbook(tp_path, data_only=True)
        ws = wb["Лист2"] if "Лист2" in wb.sheetnames else wb.active
        
        # Разбор TicketProf (колонки A-D по COL_MAP_SINGLE_TP)
        tp_items, doc_sums = parse_tp_sheet(ws, COL_MAP_SINGLE_TP)
        
        # Разбор Bars Tour (колонки H-M по COL_MAP_SINGLE_BT)
        bt_items = parse_bt_sheet(ws, COL_MAP_SINGLE_BT)
        
        wb.close()
    
    # Режим двух раздельных файлов
    else:
        # Парсим TicketProf
        wb_tp = openpyxl.load_workbook(tp_path, data_only=True)
        ws_tp = wb_tp["Лист2"] if "Лист2" in wb_tp.sheetnames else wb_tp.active
        tp_items, doc_sums = parse_tp_sheet(ws_tp, COL_MAP_DOUBLE_TP)
        wb_tp.close()
        
        # Парсим Bars Tour
        wb_bt = openpyxl.load_workbook(bt_path, data_only=True)
        ws_bt = wb_bt["Лист2"] if "Лист2" in wb_bt.sheetnames else wb_bt.active
        bt_items = parse_bt_sheet(ws_bt, COL_MAP_DOUBLE_BT)
        wb_bt.close()
        
    # --- Алгоритм распределения (аллокации) отрицательных сумм корректировок в TicketProf ---
    tp_groups = defaultdict(list)
    for item in tp_items:
        key = (item.date, item.doc)
        tp_groups[key].append(item)
            
    # Распределяем
    for key, items in tp_groups.items():
        doc_amt = doc_sums.get(key, 0.0)
        if doc_amt < 0:
            # Находим элементы с нулевой стоимостью
            zero_items = [item for item in items if item.amount == 0.0]
            non_zero_sum = sum(item.amount for item in items if item.amount != 0.0)
            remaining_to_allocate = doc_amt - non_zero_sum
            
            if zero_items and remaining_to_allocate != 0.0:
                allocated = 0.0
                for item in zero_items:
                    # Ищем соответствующий приход BT с отрицательной суммой по пересечению ID
                    bt_match = None
                    for bt in bt_items:
                        if bt.ids.intersection(item.ids) and bt.service_type == item.service_type and bt.amount < 0:
                            bt_match = bt
                            break
                    if bt_match:
                        item.allocated_amount = bt_match.amount
                        allocated += bt_match.amount
                    else:
                        # Делим поровну
                        item.allocated_amount = remaining_to_allocate / len(zero_items)
                        
                # Корректируем погрешность округления на последнем элементе
                total_allocated = sum(item.allocated_amount for item in zero_items)
                if abs(total_allocated - remaining_to_allocate) > 0.01:
                    zero_items[-1].allocated_amount += (remaining_to_allocate - total_allocated)
                    
    return tp_items, bt_items

def parse_tp_sheet(ws: Any, col_map: dict) -> Tuple[List[ServiceItem], dict]:
    """
    Разбирает лист TicketProf и возвращает список элементов и суммы документов-родителей
    """
    tp_items = []
    doc_sums = {}
    
    current_doc = None
    current_date = None
    
    for r in range(2, ws.max_row + 1):
        val_date = ws.cell(row=r, column=col_map["date"]).value
        val_doc = ws.cell(row=r, column=col_map["doc"]).value
        val_debit = ws.cell(row=r, column=col_map["debit"]).value
        val_credit = ws.cell(row=r, column=col_map["credit"]).value
        
        is_header = False
        if val_date is not None and val_doc is not None:
            if any(kw in str(val_doc) for kw in ["Продажа", "Оплата", "Возврат", "Корректировка"]):
                is_header = True
                
        if is_header:
            current_date = str(val_date).strip()
            current_doc = str(val_doc).strip()
            c_val = to_float(val_debit)
            d_val = to_float(val_credit)
            current_doc_amt = c_val if c_val is not None else (-d_val if d_val is not None else 0.0)
            doc_sums[(current_date, current_doc)] = current_doc_amt
        elif val_date is not None and val_doc is None:
            desc = str(val_date).strip()
            if not any(kw in desc for kw in ["Обороты за период", "Сальдо конечное", "Итого"]):
                c_val = to_float(val_debit)
                d_val = to_float(val_credit)
                amt = c_val if c_val is not None else (-d_val if d_val is not None else 0.0)
                tp_items.append(ServiceItem(
                    row=r,
                    date=current_date or "",
                    doc=current_doc or "",
                    desc=desc,
                    clean_desc=clean_text(desc),
                    service_type=classify_service(desc),
                    amount=amt,
                    allocated_amount=amt,
                    ids=extract_identifiers(desc),
                    source="TP"
                ))
                
    return tp_items, doc_sums

def parse_bt_sheet(ws: Any, col_map: dict) -> List[ServiceItem]:
    """
    Разбирает лист Bars Tour и возвращает список элементов
    """
    bt_items = []
    
    current_doc = None
    current_date = None
    
    for r in range(2, ws.max_row + 1):
        val_date = ws.cell(row=r, column=col_map["date"]).value
        val_doc = ws.cell(row=r, column=col_map["doc"]).value
        val_amt = ws.cell(row=r, column=col_map["amount"]).value
        val_profit = ws.cell(row=r, column=col_map["profit"]).value
        val_net = ws.cell(row=r, column=col_map["net"]).value
        
        is_header = False
        if val_date is not None and val_doc is not None:
            if any(kw in str(val_doc) for kw in ["Приход", "Оплата", "Возврат", "Принято"]):
                is_header = True
                
        if is_header:
            current_date = str(val_date).strip()
            current_doc = str(val_doc).strip()
        elif val_date is not None and val_doc is None:
            desc = str(val_date).strip()
            if not any(kw in desc for kw in ["Обороты за период", "Сальдо конечное", "Итого"]):
                j_val = to_float(val_amt)
                amt = j_val if j_val is not None else 0.0
                if amt == 0.0 and not extract_identifiers(desc):
                    continue
                    
                profit_val = to_float(val_profit)
                net_val = to_float(val_net)
                
                bt_items.append(ServiceItem(
                    row=r,
                    date=current_date or "",
                    doc=current_doc or "",
                    desc=desc,
                    clean_desc=clean_text(desc),
                    service_type=classify_service(desc),
                    amount=amt,
                    allocated_amount=amt,
                    ids=extract_identifiers(desc),
                    profit=profit_val,
                    net=net_val,
                    source="BT"
                ))
                
    # Группировка элементов Bars Tour (схлопывание строк с одинаковым описанием под одним документом)
    from collections import defaultdict
    grouped_bt = defaultdict(list)
    for item in bt_items:
        key = (item.doc, item.clean_desc, item.service_type)
        grouped_bt[key].append(item)
        
    final_bt_items = []
    for key, items in grouped_bt.items():
        if len(items) == 1:
            final_bt_items.append(items[0])
        else:
            # Сортируем элементы по модулю суммы для правильной классификации (Profit < Net < Total)
            sorted_items = sorted(items, key=lambda x: abs(x.amount))
            best_item = sorted_items[-1]  # Элемент с максимальной суммой (Total)
            
            # Объединяем все извлеченные ID
            merged_ids = set()
            for x in items:
                merged_ids.update(x.ids)
            best_item.ids = merged_ids
            
            if len(sorted_items) == 2:
                # Если 2 элемента: меньший - profit, больший - net
                best_item.profit = sorted_items[0].amount
                best_item.net = sorted_items[1].amount
                best_item.amount = sorted_items[0].amount + sorted_items[1].amount
                best_item.allocated_amount = best_item.amount
            elif len(sorted_items) >= 3:
                # Если 3 элемента: меньший - profit, средний - net, максимальный - total (amount)
                best_item.profit = sorted_items[0].amount
                best_item.net = sorted_items[1].amount
                best_item.amount = sorted_items[2].amount
                best_item.allocated_amount = best_item.amount
                
            best_item.row = items[0].row
            final_bt_items.append(best_item)
            
    return final_bt_items
