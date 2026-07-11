import openpyxl
import sys

files = [
    (r"C:\Users\Maxim\Downloads\11 июля.xlsx", "sheet_inspection_11.txt"),
    (r"C:\Users\Maxim\Downloads\12 июля.xlsx", "sheet_inspection_12.txt")
]

for file_path, out_path in files:
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(f"=== INSPECTING EXCEL FILE: {file_path} ===\n")
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            f.write(f"Sheets in workbook: {wb.sheetnames}\n\n")
            
            for name in wb.sheetnames:
                ws = wb[name]
                f.write(f"--- Sheet: {name} ---\n")
                f.write(f"Dimensions: {ws.dimensions}\n")
                f.write("First 25 rows:\n")
                for r in range(1, 40):
                    row_vals = [ws.cell(row=r, column=c).value for c in range(1, 25)]
                    row_str = " | ".join([str(val).replace('\n', ' ') if val is not None else "" for val in row_vals])
                    # Trim excess empty separators from the right
                    while row_str.endswith(" | "):
                        row_str = row_str[:-3]
                    f.write(f"Row {r:02d}: {row_str}\n")
                f.write("\n")
            wb.close()
            f.write("Inspection completed successfully.\n")
        except Exception as e:
            import traceback
            f.write(f"Error occurred: {str(e)}\n")
            traceback.print_exc(file=f)

    print("Saved inspection to", out_path)

