import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\Maxim\Downloads\13-19.07.xlsx', data_only=True)
ws = wb.active

with open('raw_hotel_rows.txt', 'w', encoding='utf-8') as f:
    f.write("=== RAW ROWS 15-20 of 13-19.07.xlsx ===\n")
    for r in range(15, 21):
        f.write(f"\nRow {r}:\n")
        for c in range(1, 12):
            v = ws.cell(r, c).value
            f.write(f"  Col {c:2d} ({openpyxl.utils.get_column_letter(c)}{r}): {repr(v)}\n")

print("Done.")
