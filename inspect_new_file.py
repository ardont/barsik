import openpyxl
import sys

file_path = r"C:\Users\Maxim\Downloads\13-19.07_сопоставлено.xlsx"
out_path = "inspect_new_file.txt"

with open(out_path, "w", encoding="utf-8") as f:
    f.write(f"=== INSPECTING EXCEL FILE: {file_path} ===\n")
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        f.write(f"Sheets in workbook: {wb.sheetnames}\n\n")
        
        for name in wb.sheetnames:
            ws = wb[name]
            f.write(f"--- Sheet: {name} ---\n")
            f.write(f"Dimensions: {ws.dimensions}\n")
            f.write("First 100 rows:\n")
            
            # Let's inspect rows
            max_r = min(ws.max_row, 300)
            for r in range(1, max_r + 1):
                row_vals = [ws.cell(row=r, column=c).value for c in range(1, 16)]
                row_str = " | ".join([str(val).replace('\n', ' ') if val is not None else "" for val in row_vals])
                # Trim excess empty separators from the right
                while row_str.endswith(" | "):
                    row_str = row_str[:-3]
                f.write(f"Row {r:03d}: {row_str}\n")
            f.write("\n")
        wb.close()
        f.write("Inspection completed successfully.\n")
    except Exception as e:
        import traceback
        f.write(f"Error occurred: {str(e)}\n")
        traceback.print_exc(file=f)

print("Saved inspection to", out_path)
