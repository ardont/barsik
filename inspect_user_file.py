import openpyxl

file_path = r"C:\Users\Maxim\Downloads\20-26_07_26.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)

out_lines = []
out_lines.append(f"Sheet names: {wb.sheetnames}")

for sname in wb.sheetnames:
    ws = wb[sname]
    out_lines.append(f"\n--- SHEET: {sname} (max_row={ws.max_row}, max_col={ws.max_column}) ---")
    
    for r in range(1, ws.max_row + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if not any(v is not None for v in row_vals):
            continue
        
        cell_info = []
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            val = cell.value
            comment = cell.comment.text if cell.comment else ""
            
            fill_color = ""
            if cell.fill and cell.fill.fgColor:
                fg = cell.fill.fgColor
                fill_color = str(fg.rgb or fg.theme or fg.indexed or "")
            
            font_color = ""
            if cell.font and cell.font.color:
                fc = cell.font.color
                font_color = str(fc.rgb or fc.theme or fc.indexed or "")
            
            if val is not None or comment or fill_color or font_color:
                info = f"C{c}: val={repr(val)}"
                if fill_color:
                    info += f" fill={fill_color}"
                if font_color:
                    info += f" font={font_color}"
                if comment:
                    info += f" comment={repr(comment)}"
                cell_info.append(info)
        
        out_lines.append(f"Row {r:3d}: " + " | ".join(cell_info))

with open("inspect_20_26_report.txt", "w", encoding="utf-8") as f:
    f.write("\n".join(out_lines))

print("Wrote inspect_20_26_report.txt")
