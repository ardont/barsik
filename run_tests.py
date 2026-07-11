# -*- coding: utf-8 -*-
"""
Автоматизированный скрипт тестирования сверки в гибридном режиме (1 или 2 файла)
"""

import os
import openpyxl
from engine.loader import load_data
from engine.matcher import match_records
from engine.calculator import calculate_reconciliation
from reports.excel_export import export_to_excel
from reports.word_export import export_to_word

def run_tests():
    print("=== ЗАПУСК ТЕСТОВ СВЕРКИ (ГИБРИДНЫЙ РЕЖИМ) ===")
    
    file_path = r"C:\Users\Maxim\Downloads\11 июля.xlsx"
    if not os.path.exists(file_path):
        print(f"Ошибка: Исходный файл не найден по пути: {file_path}")
        return
        
    # ----------------------------------------------------
    # Тест 1: Сводный монолитный файл (Legacy)
    # ----------------------------------------------------
    print("\n[Тест 1] Запуск разбора единого сводного файла...")
    tp_items_single, bt_items_single = load_data(file_path)
    print(f"Загружено из сводного: TicketProf = {len(tp_items_single)}, Bars Tour = {len(bt_items_single)}")
    
    matches_single, unmatched_tp_single, unmatched_bt_single = match_records(
        tp_items_single, bt_items_single, {}
    )
    summary_single = calculate_reconciliation(tp_items_single, bt_items_single, matches_single)
    
    print(f"Сопоставлено: {len(matches_single)}, Прибыль: {summary_single.total_profit:,.2f} руб.")
    
    # ----------------------------------------------------
    # Тест 2: Разделение на два файла (Modern)
    # ----------------------------------------------------
    print("\n[Тест 2] Разрезание сводного файла на два независимых реестра...")
    
    # Создаем временные книги
    wb_tp = openpyxl.Workbook()
    ws_tp = wb_tp.active
    ws_tp.title = "Лист2"
    
    wb_bt = openpyxl.Workbook()
    ws_bt = wb_bt.active
    ws_bt.title = "Лист2"
    
    wb_src = openpyxl.load_workbook(file_path, data_only=True)
    ws_src = wb_src["Лист2"] if "Лист2" in wb_src.sheetnames else wb_src.active
    
    for r in range(1, ws_src.max_row + 1):
        # Копируем TicketProf (колонки 1-7)
        for col_idx in range(1, 8):
            ws_tp.cell(row=r, column=col_idx, value=ws_src.cell(row=r, column=col_idx).value)
        # Копируем Bars Tour (колонки 8-13) и сдвигаем их в начало (колонки 1-6)
        for col_idx in range(8, 14):
            ws_bt.cell(row=r, column=col_idx - 7, value=ws_src.cell(row=r, column=col_idx).value)
            
    tp_temp_path = "temp_tp_sales.xlsx"
    bt_temp_path = "temp_bt_receipts.xlsx"
    
    wb_tp.save(tp_temp_path)
    wb_bt.save(bt_temp_path)
    wb_src.close()
    
    print("Временные файлы созданы. Запуск разбора двух раздельных файлов...")
    tp_items_double, bt_items_double = load_data(tp_temp_path, bt_temp_path)
    print(f"Загружено из раздельных: TicketProf = {len(tp_items_double)}, Bars Tour = {len(bt_items_double)}")
    
    # Проверки эквивалентности загруженных данных
    assert len(tp_items_double) == len(tp_items_single), "Количество элементов TicketProf в обоих режимах должно совпадать!"
    assert len(bt_items_double) == len(bt_items_single), "Количество элементов Bars Tour в обоих режимах должно совпадать!"
    
    matches_double, unmatched_tp_double, unmatched_bt_double = match_records(
        tp_items_double, bt_items_double, {}
    )
    summary_double = calculate_reconciliation(tp_items_double, bt_items_double, matches_double)
    
    print(f"Сопоставлено: {len(matches_double)}, Прибыль: {summary_double.total_profit:,.2f} руб.")
    
    # Проверки финансовой эквивалентности
    assert summary_double.total_profit == summary_single.total_profit, "Сумма прибыли в обоих режимах должна совпадать!"
    assert summary_double.total_tp_sum == summary_single.total_tp_sum, "Сумма TicketProf должна совпадать!"
    assert summary_double.total_bt_sum == summary_single.total_bt_sum, "Сумма Bars Tour должна совпадать!"
    
    # Очистка временных файлов
    os.remove(tp_temp_path)
    os.remove(bt_temp_path)
    print("Временные файлы очищены. Тест эквивалентности режимов пройден!")
    
    # ----------------------------------------------------
    # Тест 3: Экспорт отчетов
    # ----------------------------------------------------
    print("\n[Тест 3] Проверка экспорта отчетов (Excel и Word)...")
    excel_out = "test_hybrid_report.xlsx"
    export_to_excel(tp_items_double, bt_items_double, matches_double, unmatched_tp_double, unmatched_bt_double, summary_double, excel_out)
    assert os.path.exists(excel_out), "Файл Excel отчета должен быть создан"
    print(f"Excel отчет успешно сохранен в: {os.path.abspath(excel_out)}")
    os.remove(excel_out)
    
    word_out = "test_hybrid_act.docx"
    export_to_word(unmatched_tp_double, unmatched_bt_double, summary_double, word_out)
    assert os.path.exists(word_out), "Файл Word Акта сверки должен быть создан"
    print(f"Word Акт сверки успешно сохранен в: {os.path.abspath(word_out)}")
    os.remove(word_out)
    
    print("\n=== ВСЕ ТЕСТЫ ГИБРИДНОГО РЕЖИМА ПРОЙДЕНЫ УСПЕШНО! ===")

if __name__ == "__main__":
    try:
        run_tests()
    except Exception as e:
        import traceback
        print(f"\nОшибка во время выполнения тестов: {e}")
        traceback.print_exc()
