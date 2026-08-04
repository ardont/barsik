import openpyxl
import re

def extract_identifiers_v2(text: str):
    if not text:
        return set()
    
    found = set()
    
    # 1. Авиабилеты: 3-4 символа + дефис + 10 цифр (e.g. 555-2396327717, 15К-6111740577)
    air_tickets = re.findall(r'\b[0-9А-ЯA-Zа-яa-z]{2,4}-\d{10}\b', text)
    found.update(air_tickets)
    
    # 2. Ж/Д билеты: ЭЖБ-14цифр или просто 14 цифр подряд (e.g. ЭЖБ-75014939788965, 75014939788965)
    # Ищем с префиксом ЭЖБ или без
    rail_tickets_with_prefix = re.findall(r'\b[А-ЯA-Zа-яa-z]{3}-\d{14}\b', text)
    found.update(rail_tickets_with_prefix)
    
    rail_tickets_digits = re.findall(r'\b75\d{12}\b', text)
    found.update(rail_tickets_digits)
    
    # 3. Заказы (отели и др.): 7-10 цифр после Заказ / №
    orders = re.findall(r'(?:Заказ|заказ|заказа|№)\s*(\d{7,10})\b', text)
    found.update(orders)
    
    # 4. Если есть № в скобках типа (ФИО: № ЭЖБ-75014939788965)
    bracket_ids = re.findall(r'№\s*([0-9А-ЯA-Zа-яa-z\-]{6,20})', text)
    for b_id in bracket_ids:
        b_id_clean = b_id.strip()
        if len(b_id_clean) >= 7:
            found.add(b_id_clean)
            
    return found

wb = openpyxl.load_workbook(r'C:\Users\Maxim\Downloads\13-19.07.xlsx', data_only=True)
ws = wb.active

print("=== TESTING EXTRACT IDENTIFIERS V2 ON 13-19.07.xlsx ===")
with open('test_regex_out.txt', 'w', encoding='utf-8') as out:
    for r in range(1, ws.max_row + 1):
        tp_desc = ws.cell(r, 1).value
        bt_desc = ws.cell(r, 8).value
        
        if tp_desc and not any(kw in str(tp_desc) for kw in ["По данным", "Дата", "Обороты", "Сальдо"]):
            ids = extract_identifiers_v2(str(tp_desc))
            out.write(f"Row {r:2d} TP IDs: {ids} | desc={tp_desc}\n")
        if bt_desc and not any(kw in str(bt_desc) for kw in ["По данным", "Дата", "Обороты", "Сальдо"]):
            ids = extract_identifiers_v2(str(bt_desc))
            out.write(f"Row {r:2d} BT IDs: {ids} | desc={bt_desc}\n")

print("Done testing.")
