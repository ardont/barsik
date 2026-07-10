import openpyxl
import re

file_path = r"C:\Users\Maxim\Downloads\08.07_1.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb["Лист2"]

def extract_id(text):
    if not text:
        return None, None
    ticket_match = re.search(r'\b\d{3}-\d{10}\b', text)
    if ticket_match:
        return "ticket", ticket_match.group(0)
    order_match = re.search(r'(?:Заказ|заказ)\s*(?:№)?\s*(\d+)', text)
    if order_match:
        return "order", order_match.group(1)
    return None, None

print("=== PARSING TICKETPROF ===")
tp_items = []
current_doc = None
current_date = None

for r in range(2, ws.max_row + 1):
    val_a = ws.cell(row=r, column=1).value
    val_b = ws.cell(row=r, column=2).value
    val_c = ws.cell(row=r, column=3).value
    val_d = ws.cell(row=r, column=4).value # Credit column
    
    # We identify a document header if:
    # 1. val_a is a date/string AND val_b has a document name (e.g. contains "Продажа", "Оплата", "Возврат", "Корректировка")
    # OR 2. it contains "Обороты за период" or "Сальдо"
    is_header = False
    if val_a is not None and val_b is not None:
        if any(kw in str(val_b) for kw in ["Продажа", "Оплата", "Возврат", "Корректировка"]):
            is_header = True
            
    if is_header:
        current_date = val_a
        current_doc = val_b
        print(f"Header: Row {r} | Date: {val_a} | Doc: {val_b} | Debit: {val_c} | Credit: {val_d}")
    elif val_a is not None and val_b is None:
        # This is a sub-item description row!
        desc = str(val_a).strip()
        id_type, id_val = extract_id(desc)
        amt = val_c if val_c is not None else ( -val_d if val_d is not None else 0.0 )
        tp_items.append({
            "row": r,
            "date": current_date,
            "doc": current_doc,
            "desc": desc,
            "amount": amt,
            "id_type": id_type,
            "id_val": id_val
        })
        print(f"  Sub-item: Row {r} | Amt: {amt} | ID: {id_val} | Desc: {desc[:50]}...")

wb.close()
