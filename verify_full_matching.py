import openpyxl
import re
import os
from collections import defaultdict

file_path = r"C:\Users\Maxim\Downloads\08.07_1.xlsx"

def classify_service(text):
    t = text.lower()
    if any(kw in t for kw in ["проживание", "отель", "г-ца", "гостиница"]):
        return "Hotel"
    if any(kw in t for kw in ["выбор места", "место", "mco"]):
        return "Seat"
    if any(kw in t for kw in ["услуга ито", "сервисный сбор", "ито"]):
        return "Fee"
    if any(kw in t for kw in ["штраф", "удержания", "удержание"]):
        return "Penalty"
    if any(kw in t for kw in ["авиабилет", "билет", "возврат авиабилета"]):
        return "Flight"
    return "Other"

def extract_identifiers(text):
    if not text:
        return set()
    tickets = re.findall(r'\b\d{3}-\d{10}\b', text)
    orders = re.findall(r'(?:Заказ|заказ|заказа|№)\s*(\d{7,10})\b', text)
    return set(tickets + orders)

wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb["Лист2"]

# Parse TicketProf items
tp_items = []
current_doc_tp = None
current_date_tp = None
current_doc_tp_row = None
current_doc_tp_amt = 0.0

for r in range(2, ws.max_row + 1):
    val_a = ws.cell(row=r, column=1).value
    val_b = ws.cell(row=r, column=2).value
    val_c = ws.cell(row=r, column=3).value
    val_d = ws.cell(row=r, column=4).value
    
    is_header = False
    if val_a is not None and val_b is not None:
        if any(kw in str(val_b) for kw in ["Продажа", "Оплата", "Возврат", "Корректировка"]):
            is_header = True
            
    if is_header:
        current_date_tp = val_a
        current_doc_tp = val_b
        current_doc_tp_row = r
        current_doc_tp_amt = val_c if val_c is not None else (-val_d if val_d is not None else 0.0)
    elif val_a is not None and val_b is None:
        desc = str(val_a).strip()
        if any(kw in desc for kw in ["Обороты за период", "Сальдо конечное", "Итого"]):
            continue
        amt = val_c if val_c is not None else (-val_d if val_d is not None else 0.0)
        tp_items.append({
            "row": r,
            "date": current_date_tp,
            "doc": current_doc_tp,
            "doc_row": current_doc_tp_row,
            "doc_amount": current_doc_tp_amt,
            "desc": desc,
            "service_type": classify_service(desc),
            "amount": amt,
            "ids": extract_identifiers(desc),
            "allocated_amount": amt
        })

# Parse Bars Tour items
bt_items = []
current_doc_bt = None
current_date_bt = None
current_doc_bt_row = None
current_doc_bt_amt = 0.0

for r in range(2, ws.max_row + 1):
    val_h = ws.cell(row=r, column=8).value
    val_i = ws.cell(row=r, column=9).value
    val_j = ws.cell(row=r, column=10).value
    val_l = ws.cell(row=r, column=12).value
    val_m = ws.cell(row=r, column=13).value
    
    is_header = False
    if val_h is not None and val_i is not None:
        if any(kw in str(val_i) for kw in ["Приход", "Оплата", "Возврат", "Принято"]):
            is_header = True
            
    if is_header:
        current_date_bt = val_h
        current_doc_bt = val_i
        current_doc_bt_row = r
        current_doc_bt_amt = val_j if val_j is not None else 0.0
    elif val_h is not None and val_i is None:
        desc = str(val_h).strip()
        if any(kw in desc for kw in ["Обороты за период", "Сальдо конечное", "Итого"]):
            continue
        amt = val_j if val_j is not None else 0.0
        if amt == 0.0 and not extract_identifiers(desc):
            continue
        bt_items.append({
            "row": r,
            "date": current_date_bt,
            "doc": current_doc_bt,
            "doc_row": current_doc_bt_row,
            "doc_amount": current_doc_bt_amt,
            "desc": desc,
            "service_type": classify_service(desc),
            "amount": amt,
            "profit": val_l,
            "net": val_m,
            "ids": extract_identifiers(desc)
        })

wb.close()

# ALLOCATION ENGINE: allocate 0-amount TP items if the header amount is negative
# Group TP items by document row
tp_by_doc = defaultdict(list)
for item in tp_items:
    tp_by_doc[item["doc_row"]].append(item)

for doc_row, items in tp_by_doc.items():
    doc_amt = items[0]["doc_amount"]
    if doc_amt < 0:
        # Check if we have 0-amount items
        zero_items = [item for item in items if item["amount"] == 0.0]
        non_zero_sum = sum(item["amount"] for item in items if item["amount"] != 0.0)
        remaining_to_allocate = doc_amt - non_zero_sum
        
        if zero_items and remaining_to_allocate != 0.0:
            # Let's allocate by looking up corresponding BT return values
            allocated = 0.0
            for item in zero_items:
                # Find matching BT item under same IDs and service type with negative amount
                bt_match = None
                for bt in bt_items:
                    if bt["ids"].intersection(item["ids"]) and bt["service_type"] == item["service_type"] and bt["amount"] < 0:
                        bt_match = bt
                        break
                if bt_match:
                    item["allocated_amount"] = bt_match["amount"]
                    allocated += bt_match["amount"]
                else:
                    # Fallback: divide equally
                    item["allocated_amount"] = remaining_to_allocate / len(zero_items)
            
            # If the allocation doesn't match the remaining, adjust the last item
            total_allocated = sum(item["allocated_amount"] for item in zero_items)
            if abs(total_allocated - remaining_to_allocate) > 0.01:
                zero_items[-1]["allocated_amount"] += (remaining_to_allocate - total_allocated)

# Now, we group both streams by ID
tp_by_id = defaultdict(list)
for item in tp_items:
    for ident in item["ids"]:
        tp_by_id[ident].append(item)

bt_by_id = defaultdict(list)
for item in bt_items:
    for ident in item["ids"]:
        bt_by_id[ident].append(item)

all_ids = set(tp_by_id.keys()).union(set(bt_by_id.keys()))

with open("matching_report.txt", "w", encoding="utf-8") as f:
    f.write("=== DETAILED ID-LEVEL MATCHING AND FINANCIALS ===\n\n")
    
    unmatched_tp_total = 0.0
    unmatched_bt_total = 0.0
    matched_tp_total = 0.0
    matched_bt_total = 0.0
    profit_total = 0.0
    
    f.write(f"{'ID':<15} | {'Type':<8} | {'TP Sum':<12} | {'BT Sum':<12} | {'Profit':<12} | {'Status':<15}\n")
    f.write("-" * 80 + "\n")
    
    for ident in sorted(all_ids):
        tp_list = tp_by_id[ident]
        bt_list = bt_by_id[ident]
        
        # De-duplicate items that appear multiple times under multiple IDs
        # To get the real financial sums, we sum the unique items by their row number
        tp_unique = {item["row"]: item for item in tp_list}.values()
        bt_unique = {item["row"]: item for item in bt_list}.values()
        
        tp_sum = sum(item["allocated_amount"] for item in tp_unique)
        bt_sum = sum(item["amount"] for item in bt_unique)
        
        # Status calculation
        status = "Match"
        profit = 0.0
        
        if not tp_unique:
            status = "Only BT"
            unmatched_bt_total += bt_sum
        elif not bt_unique:
            status = "Only TP"
            unmatched_tp_total += tp_sum
        else:
            # Both present
            matched_tp_total += tp_sum
            matched_bt_total += bt_sum
            # Profit formula: BT - TP
            # Wait! If the service type is flight, BT sum includes the 10% fee. So profit = BT - TP
            # For returns: BT is negative, TP is negative.
            # Example: BT = -26291, TP = -26021. Profit = BT - TP = -26291 - (-26021) = -270 (which is loss/fee)
            # In general: profit = BT_sum - TP_sum
            profit = bt_sum - tp_sum
            profit_total += profit
            
            if abs(profit) > 0.01:
                status = f"Diff: {profit:+.2f}"
            else:
                status = "Exact Match"
                
        f.write(f"{ident:<15} | {list(tp_unique)[0]['service_type'] if tp_unique else list(bt_unique)[0]['service_type']:<8} | {tp_sum:<12.2f} | {bt_sum:<12.2f} | {profit:<12.2f} | {status:<15}\n")

    f.write("\n=== SUMMARY ===\n")
    f.write(f"Matched TicketProf Total: {matched_tp_total:,.2f} руб.\n")
    f.write(f"Matched Bars Tour Total: {matched_bt_total:,.2f} руб.\n")
    f.write(f"Total Profit Calculated: {profit_total:,.2f} руб.\n")
    f.write(f"Unmatched TicketProf Total: {unmatched_tp_total:,.2f} руб.\n")
    f.write(f"Unmatched Bars Tour Total: {unmatched_bt_total:,.2f} руб.\n")

print("Saved report to matching_report.txt")
