import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\Maxim\Downloads\13-19.07.xlsx', data_only=True)
ws = wb.active

with open('hotel_bt_cols.txt', 'w', encoding='utf-8') as f:
    f.write("=== BARS TOUR COLUMNS 8 to 13 IN 13-19.07.xlsx ===\n")
    for r in range(1, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(8, 14)]
        if any(v is not None for v in vals):
            f.write(f"Row {r:2d}: H={repr(vals[0])} | I={repr(vals[1])} | J={repr(vals[2])} | K={repr(vals[3])} | L={repr(vals[4])} | M={repr(vals[5])}\n")

print("Done.")
