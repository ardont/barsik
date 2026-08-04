import openpyxl
import re
from engine.loader import to_float

def clean_text(text: str) -> str:
    if not text:
        return ""
    t = text.lower()
    t = re.sub(r'\b\d{2}\.\d{2}\.(?:\d{4}|\d{2})\b', '', t)
    t = re.sub(r'\bс\s+\d{2}\.\d{2}\s+по\s+\d{2}\.\d{2}\b', '', t)
    t = re.sub(r'\bс\s+\d{2}\.\d{2}\.\d{2,4}\s+по\s+\d{2}\.\d{2}\.\d{2,4}\b', '', t)
    t = re.sub(r'\([a-z0-9\s\/\-\:\.\,№а-я]+\)', '', t)
    t = re.sub(r'\b[a-z]+/[a-z]+\b', '', t)
    t = re.sub(r'\b[0-9а-яa-z]{2,4}-\d{10,14}\b', '', t)
    t = re.sub(r'\b\d{13,14}\b', '', t)
    t = re.sub(r'\b(?:заказ|заказа|№)?\s*\d+\b', '', t)
    t = re.sub(r'[^\w\s]', ' ', t)
    t = re.sub(r'\s+', ' ', t).strip()
    return t

def classify_service(text: str) -> str:
    if not text:
        return "Other"
    t = text.lower()
    if any(kw in t for kw in ["проживание", "отель", "г-ца", "гостиница"]):
        return "Hotel"
    if any(kw in t for kw in ["выбор места", "место", "mco"]):
        return "Seat"
    if any(kw in t for kw in ["услуга ито", "сервисный сбор", "ито"]):
        return "Fee"
    if any(kw in t for kw in ["штраф", "удержания", "удержание"]):
        return "Penalty"
    if any(kw in t for kw in ["авиабилет", "билет", "ж.д.", "перелет", "поезд"]):
        return "Flight"
    return "Other"

def extract_identifiers(text: str):
    if not text:
        return set()
    found = set()
    
    # 1. Авиабилеты: 2-4 символа + дефис + 10 цифр (555-2396327717, 15К-6111740577)
    air_tickets = re.findall(r'\b[0-9А-ЯA-Zа-яa-z]{2,4}-\d{10}\b', text)
    for t in air_tickets:
        found.add(t)
    
    # 2. Ж/Д билеты: ЭЖБ-14 цифр или 14 цифр подряд (ЭЖБ-75014939788965)
    rail_tickets = re.findall(r'\b[А-ЯA-Zа-яa-z]{3}-\d{14}\b', text)
    for t in rail_tickets:
        found.add(t)
        found.add(t.split('-')[-1])
        
    rail_digits = re.findall(r'\b75\d{12}\b', text)
    for d in rail_digits:
        found.add(d)
        found.add(f"ЭЖБ-{d}")
        
    # 3. Заказы отелей: 7-10 цифр после Заказ/№
    orders = re.findall(r'(?:Заказ|заказ|заказа|№)\s*(\d{7,10})\b', text)
    for o in orders:
        found.add(o)
        
    return found

def get_primary_id(ids) -> str:
    if not ids:
        return "N/A"
    with_hyphen = [i for i in ids if '-' in i]
    if with_hyphen:
        return with_hyphen[0]
    sorted_ids = sorted(list(ids), key=lambda x: len(x), reverse=True)
    return sorted_ids[0]

wb = openpyxl.load_workbook(r'C:\Users\Maxim\Downloads\13-19.07.xlsx', data_only=True)
ws = wb.active

# Parse TP (Cols 1-4)
tp_items = []
current_doc = None
current_date = None

for r in range(2, ws.max_row + 1):
    v_date = ws.cell(r, 1).value
    v_doc = ws.cell(r, 2).value
    v_deb = ws.cell(r, 3).value
    v_cred = ws.cell(r, 4).value
    
    if v_date is not None and v_doc is not None:
        if any(kw in str(v_doc) for kw in ["Продажа", "Оплата", "Возврат", "Корректировка"]):
            current_date = str(v_date).strip()
            current_doc = str(v_doc).strip()
    elif v_date is not None and v_doc is None:
        desc = str(v_date).strip()
        if not any(kw in desc for kw in ["Обороты за период", "Сальдо конечное", "Итого"]):
            c_val = to_float(v_deb)
            d_val = to_float(v_cred)
            amt = c_val if c_val is not None else (-d_val if d_val is not None else 0.0)
            tp_items.append({
                'row': r, 'date': current_date, 'doc': current_doc, 'desc': desc,
                'clean_desc': clean_text(desc), 'service_type': classify_service(desc),
                'amount': amt, 'ids': extract_identifiers(desc), 'matched': False
            })

# Parse BT (Cols 8-11)
bt_items = []
current_doc = None
current_date = None

for r in range(2, ws.max_row + 1):
    v_date = ws.cell(r, 8).value
    v_doc = ws.cell(r, 9).value
    v_amt = ws.cell(r, 10).value
    
    if v_date is not None and v_doc is not None:
        if any(kw in str(v_doc) for kw in ["Приход", "Оплата", "Возврат", "Принято"]):
            current_date = str(v_date).strip()
            current_doc = str(v_doc).strip()
    elif v_date is not None and v_doc is None:
        desc = str(v_date).strip()
        if not any(kw in desc for kw in ["Обороты за период", "Сальдо конечное", "Итого"]):
            amt = to_float(v_amt) or 0.0
            if amt == 0.0 and not extract_identifiers(desc):
                continue
            bt_items.append({
                'row': r, 'date': current_date, 'doc': current_doc, 'desc': desc,
                'clean_desc': clean_text(desc), 'service_type': classify_service(desc),
                'amount': amt, 'ids': extract_identifiers(desc), 'matched': False
            })

matches = []

# Step 1: ID Match
for tp in tp_items:
    if not tp['matched'] and tp['ids']:
        bt_candidates = [b for b in bt_items if not b['matched'] and b['service_type'] == tp['service_type'] and b['ids'].intersection(tp['ids']) and ((tp['amount'] >= 0 and b['amount'] >= 0) or (tp['amount'] < 0 and b['amount'] < 0))]
        if bt_candidates:
            best_bt = min(bt_candidates, key=lambda b: abs(b['amount'] - tp['amount']))
            tp['matched'] = True
            best_bt['matched'] = True
            matches.append((tp, best_bt, "По ID билета/заказа"))

# Step 2: Exact Name Match
for tp in tp_items:
    if not tp['matched']:
        bt_candidates = [b for b in bt_items if not b['matched'] and b['service_type'] == tp['service_type'] and b['clean_desc'] == tp['clean_desc'] and ((tp['amount'] >= 0 and b['amount'] >= 0) or (tp['amount'] < 0 and b['amount'] < 0))]
        if bt_candidates:
            best_bt = min(bt_candidates, key=lambda b: abs(b['amount'] - tp['amount']))
            tp['matched'] = True
            best_bt['matched'] = True
            matches.append((tp, best_bt, "Точное имя"))

unmatched_tp = [tp for tp in tp_items if not tp['matched']]
unmatched_bt = [bt for bt in bt_items if not bt['matched']]

with open('v2_report.txt', 'w', encoding='utf-8') as out:
    out.write(f"TOTAL TP: {len(tp_items)}, TOTAL BT: {len(bt_items)}\n")
    out.write(f"MATCHES: {len(matches)}\n\n")
    for tp, bt, method in matches:
        pid = get_primary_id(tp['ids'].union(bt['ids']))
        out.write(f"MATCH [{method:20s}] ID: {pid:20s} | TP(row {tp['row']:2d}, amt={tp['amount']:8.1f}): {tp['desc'][:50]} <-> BT(row {bt['row']:2d}, amt={bt['amount']:8.1f}): {bt['desc'][:50]}\n")
        
    out.write(f"\nUNMATCHED TP ({len(unmatched_tp)}):\n")
    for tp in unmatched_tp:
        out.write(f"  TP [row {tp['row']:2d}] amt={tp['amount']:8.1f} | ids={tp['ids']} | {tp['desc']}\n")
        
    out.write(f"\nUNMATCHED BT ({len(unmatched_bt)}):\n")
    for bt in unmatched_bt:
        out.write(f"  BT [row {bt['row']:2d}] amt={bt['amount']:8.1f} | ids={bt['ids']} | {bt['desc']}\n")

print("Report created successfully.")
