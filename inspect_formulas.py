import openpyxl

file_path = r"C:\Users\Maxim\Downloads\08.07_1.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)
ws = wb["Лист2"]

rows_to_check = [5, 15, 105, 107, 109, 111]
print("Row | Col H | Col J | Col L (Formula) | Col M (Formula)")
print("-" * 80)
for r in rows_to_check:
    desc = ws.cell(row=r, column=8).value
    amt = ws.cell(row=r, column=10).value
    f_l = ws.cell(row=r, column=12).value
    f_m = ws.cell(row=r, column=13).value
    desc_trunc = str(desc)[:25] if desc else ""
    print(f"{r:3d} | {desc_trunc:<25} | {amt} | {f_l} | {f_m}")
wb.close()
