import openpyxl

wb = openpyxl.load_workbook(r"C:\Users\Maxim\Downloads\20-26_07_26.xlsx", data_only=True)
ws = wb.active

for r in range(65, 75):
    h = ws.cell(r, 8).value
    i = ws.cell(r, 9).value
    j = ws.cell(r, 10).value
    k = ws.cell(r, 11).value
    print(f"Row {r:2d} | Col8 (H): {repr(h)} | Col9 (I): {repr(i)} | Col10 (J): {repr(j)} | Col11 (K): {repr(k)}")
