import openpyxl

def scan_styles_and_text(path):
    wb = openpyxl.load_workbook(path, data_only=False)
    print("==================================================")
    print("FILE:", path)
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        print(f"\n--- Sheet: {sheetname} ---")
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(r, c)
                val = cell.value
                fill = cell.fill.start_color.rgb if cell.fill and cell.fill.fill_type else None
                font = cell.font.color.rgb if cell.font and cell.font.color else None
                comment = cell.comment.text if cell.comment else None
                
                # Check for noticeable colors (not standard white/black/transparent/default headers)
                # Let's print any non-empty cell that might be interesting or colored
                if fill and fill not in ['00000000', 'FFFFFFFF', '00FFFFFF']:
                    print(f"Row {r:2d}, Col {c:2d} ({cell.coordinate}): val={repr(val)} | fill={fill} | font={font} | comment={repr(comment)}")
                elif comment:
                    print(f"Row {r:2d}, Col {c:2d} ({cell.coordinate}) COMMENT: val={repr(val)} | comment={repr(comment)}")
                elif font and font not in ['00000000', 'FF000000', 'FFFFFFFF']:
                    print(f"Row {r:2d}, Col {c:2d} ({cell.coordinate}) FONT COLOR: val={repr(val)} | fill={fill} | font={font}")

scan_styles_and_text(r'C:\Users\Maxim\Downloads\13-19.07.xlsx')
scan_styles_and_text(r'C:\Users\Maxim\Downloads\13-19.07_сопоставлено.xlsx')
