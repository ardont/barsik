import openpyxl

wb = openpyxl.load_workbook(r"C:\Users\Maxim\Downloads\20-26_07_26.xlsx", data_only=True)
ws = wb.active

for r in range(70, 85):
    c8 = ws.cell(r, 8).value
    c9 = ws.cell(r, 9).value
    c10 = ws.cell(r, 10).value
    c11 = ws.cell(r, 11).value
    print(f"Row {r:2d} | Col8: {repr(c8)} | Col9: {repr(c9)} | Col10: {repr(c10)} | Col11: {repr(c11)}")
